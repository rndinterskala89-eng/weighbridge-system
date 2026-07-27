# dashboard/views.py
from io import BytesIO
from datetime import datetime, timedelta
from decimal import Decimal
from django.views.decorators.csrf import csrf_protect
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.models import User
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from django.db.models import Count, Sum, Avg, Max, Min, Q
from django.core.paginator import Paginator
from django.conf import settings
from django.views.decorators.csrf import csrf_exempt
from django.db import connection
from django.core.files.base import ContentFile

import requests
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
from django.core.cache import cache

from .models import *
from .forms import *

import json
import uuid
import csv
import os
import base64
import logging


# Setup logging
logger = logging.getLogger(__name__)

# ============================================
# IMPORT UNTUK PDF
# ============================================
try:
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    import io
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False
    print("⚠️ ReportLab not installed. Run: pip install reportlab")

# ============================================
# PRINT TICKET - DETEKSI PRINTER OTOMATIS
# ============================================

# ✅ IMPORT UNTUK WINDOWS PRINTER (DENGAN TRY-EXCEPT)
try:
    import win32print
    import win32ui
    WIN32_AVAILABLE = True
except ImportError:
    WIN32_AVAILABLE = False
    win32print = None
    win32ui = None
    print("⚠️ pywin32 not installed. Windows printer features disabled.")

# ============================================
# FUNGSI DETEKSI PRINTER
# ============================================

def detect_available_printer():
    """Deteksi printer yang tersedia (LX-310 atau TM-U220D)"""
    try:
        # Coba deteksi melalui QZ Tray
        try:
            response = requests.get('http://localhost:8182/api/printer/list', timeout=3)
            if response.status_code == 200:
                printers = response.json()
                for printer in printers:
                    name = printer.get('name', '')
                    if 'LX-310' in name or 'LQ' in name or 'Dot Matrix' in name:
                        return name
                    elif 'TM-U220' in name or 'TMU220' in name or 'EPSON TM' in name:
                        return name
                if printers:
                    return printers[0].get('name', '')
        except:
            pass
        
        # Coba deteksi melalui Windows (jika di Windows dan win32print tersedia)
        if WIN32_AVAILABLE and os.name == 'nt':
            try:
                if win32print is not None:
                    printers = win32print.EnumPrinters(
                        win32print.PRINTER_ENUM_LOCAL | win32print.PRINTER_ENUM_CONNECTIONS
                    )
                    for printer in printers:
                        name = printer[2]
                        if 'LX-310' in name or 'LQ' in name:
                            return name
                        elif 'TM-U220' in name or 'TMU220' in name:
                            return name
                    if printers:
                        return printers[0][2]
            except Exception as e:
                logger.warning(f"Windows printer detection error: {e}")
        
        # Default
        return 'EPSON LX-310'
        
    except Exception as e:
        logger.error(f"Detect printer error: {e}")
        return None


def detect_printer_type(printer_name):
    """Deteksi jenis printer berdasarkan nama"""
    if not printer_name:
        return 'unknown'
    printer_name = printer_name.upper()
    if 'LX-310' in printer_name or 'LQ' in printer_name or 'DOT MATRIX' in printer_name:
        return 'dotmatrix'
    elif 'TM-U220' in printer_name or 'TMU220' in printer_name or 'EPSON TM' in printer_name:
        return 'thermal'
    return 'unknown'


def center_text(text, width=80):
    """
    Center text untuk print dengan lebar tertentu
    Default width = 80 untuk LX-310
    """
    text = str(text)
    if len(text) >= width:
        return text[:width]
    # Hitung padding kiri dan kanan
    left_padding = (width - len(text)) // 2
    right_padding = width - len(text) - left_padding
    return ' ' * left_padding + text + ' ' * right_padding


def send_to_printer(printer_name, ticket, printer_type='dotmatrix'):
    """Kirim data ke printer - HANYA 1 KALI"""
    try:
        encoding = 'cp437'
        if printer_type == 'thermal':
            encoding = 'cp437'
        
        # QZ Tray
        try:
            payload = {
                'printer': printer_name,
                'data': ticket,
                'encoding': encoding,
                'copies': 1,
                'options': {
                    'copies': 1
                }
            }
            response = requests.post('http://localhost:8182/api/print', 
                                    json=payload, timeout=5)
            if response.status_code == 200:
                return {'status': 'success', 'message': 'Printed via QZ Tray'}
        except requests.exceptions.RequestException as e:
            logger.warning(f"QZ Tray print error: {e}")
        
        # Windows API
        if WIN32_AVAILABLE and os.name == 'nt':
            try:
                if win32print is not None:
                    hprinter = win32print.OpenPrinter(printer_name)
                    try:
                        printer_info = win32print.GetPrinter(hprinter, 2)
                        devmode = printer_info['pDevMode']
                        if devmode:
                            devmode.Copies = 1
                            win32print.SetPrinter(hprinter, 2, printer_info, 0)
                        
                        job_id = win32print.StartDocPrinter(hprinter, 1, ('Ticket', None, 'RAW'))
                        try:
                            win32print.StartPagePrinter(hprinter)
                            win32print.WritePrinter(hprinter, ticket.encode(encoding))
                            win32print.EndPagePrinter(hprinter)
                        finally:
                            win32print.EndDocPrinter(hprinter)
                    finally:
                        win32print.ClosePrinter(hprinter)
                    
                    return {'status': 'success', 'message': 'Printed via Windows API'}
            except Exception as e:
                logger.error(f"Windows print error: {e}")
        
        return {
            'status': 'error', 
            'message': 'Tidak ada metode print yang tersedia. Pastikan QZ Tray terinstall atau printer terhubung.'
        }
        
    except Exception as e:
        logger.error(f"Send to printer error: {e}")
        return {'status': 'error', 'message': str(e)}

# ============================================
# FUNGSI FORMAT TICKET
# ============================================

def format_ticket_dotmatrix(no_transaksi, barang, barang_id, barang_lot,
                           kustomer, kustomer_id, supplier, supplier_id,
                           weight, keterangan, operator,
                           company_name, company_address, company_phone, company_email):
    """
    Format LX-310 dengan spasi antara Kustomer dan Supplier (VERSI PADAT)
    """
    from datetime import datetime
    now = datetime.now()
    WIDTH = 10
    LINE = '=' * WIDTH
    THIN = '-' * WIDTH
    
    tanggal = now.strftime('%d-%m-%Y')
    waktu = now.strftime('%H:%M:%S')
    
    # ============================================
    # TICKET PADAT (TANPA JARAK BERLEBIH)
    # ============================================
    ticket = f"""{LINE}
{center_text(company_name, WIDTH)}
{center_text(company_address[:40], WIDTH)}
{center_text(f'TELP: {company_phone}', WIDTH)}
{LINE}
{center_text('WEIGHING IoT TICKET', WIDTH)}
{center_text(f'{tanggal}  {waktu}', WIDTH)}
{THIN}
{'No. Transaksi':<5} : {no_transaksi}
{'Nama Barang':<5} : {barang[:30]}
{'ID Barang':<5} : {barang_id}
{'Lot':<5} : {barang_lot}
{THIN}
{'Kustomer':<5} : {kustomer[:30]}
{'ID Kustomer':<5} : {kustomer_id}
{THIN}
{'Supplier':<5} : {supplier[:30]}
{'ID Supplier':<5} : {supplier_id}
{THIN}
{center_text(f'BERAT : {weight} kg', WIDTH)}
{THIN}
{'Keterangan':<5} : {keterangan[:30]}
{'Operator':<5} : {operator[:20]}
{LINE}
{center_text('TERIMA KASIH', WIDTH)}
{center_text('FDA 21 CFR Part 11', WIDTH)}
{LINE}"""
    return ticket

   # Konfigurasi Printer TM-U220D
def format_ticket_thermal(no_transaksi, barang, barang_id, barang_lot,
                         kustomer, kustomer_id, supplier, supplier_id,
                         weight, keterangan, operator,
                         company_name, company_address, company_phone, company_email):
    """Format untuk printer TM-U220D (Thermal, 58mm / 32 kolom) - 1 HALAMAN"""
    now = datetime.now()
    line = '=' * 32
    thin = '-' * 32
    
    # Singkat nama perusahaan untuk thermal
    company_short = company_name[:20] if len(company_name) > 20 else company_name
    
    ticket = f"""
{line}
{center_text(company_short, 32)}
{center_text('WEIGHING TICKET', 32)}
{thin}
No   : {no_transaksi}
Tgl  : {now.strftime('%d-%m-%Y %H:%M')}
{thin}
Brg  : {barang[:18]}
ID   : {barang_id}
Lot  : {barang_lot}
{thin}
Kust : {kustomer[:18]}
ID   : {kustomer_id}
{thin}
Supp : {supplier[:18]}
ID   : {supplier_id}
{thin}
{center_text(f'BERAT: {weight} kg', 32)}
{thin}
Ket  : {keterangan[:18]}
Opr  : {operator[:15]}
{line}
{center_text('THANK YOU', 32)}
{center_text('FDA 21 CFR Part 11', 32)}
{line}
"""
    return ticket


# ============================================
# VIEW PRINT TICKET
# ============================================

@login_required
@csrf_exempt
def print_ticket(request):
    """Print ticket dengan deteksi printer otomatis (LX-310 / TM-U220D)"""
    if request.method != 'POST':
        return JsonResponse({
            'status': 'error',
            'message': 'Method not allowed'
        }, status=405)
    
    try:
        data = json.loads(request.body)
        
        # Ambil data dari request
        no_transaksi = data.get('no_transaksi', '')
        barang = data.get('barang', '')
        barang_id = data.get('barang_id', '')
        barang_lot = data.get('barang_lot', '')
        kustomer = data.get('kustomer', '')
        kustomer_id = data.get('kustomer_id', '')
        supplier = data.get('supplier', '')
        supplier_id = data.get('supplier_id', '')
        weight = data.get('weight', '0')
        keterangan = data.get('keterangan', '')
        operator = request.user.get_full_name() or request.user.username
        
        # Ambil company profile
        company = CompanyProfile.objects.first()
        company_name = company.name if company else 'PT Interskala Mandiri Indonesia'
        company_address = company.address if company else 'Green Sedayu Biz Park Jl. Daan Mogot KM. 18, Kalideres, Jakarta Barat'
        company_phone = company.phone if company else '(021) 2252-2992'
        company_email = company.email if company else 'sales@interskala.com'
        
        # Deteksi printer yang tersedia
        printer_name = detect_available_printer()
        
        if not printer_name:
            return JsonResponse({
                'status': 'error',
                'message': 'Tidak ada printer yang terdeteksi. Pastikan printer terhubung dan driver terinstall.'
            }, status=404)
        
        # Tentukan jenis printer
        printer_type = detect_printer_type(printer_name)
        
        # Format ticket sesuai jenis printer
        if printer_type == 'thermal' or 'TM-U220' in printer_name or 'TMU220' in printer_name:
            ticket = format_ticket_thermal(
                no_transaksi, barang, barang_id, barang_lot,
                kustomer, kustomer_id, supplier, supplier_id,
                weight, keterangan, operator,
                company_name, company_address, company_phone, company_email
            )
            printer_type_display = 'Epson TM-U220D (Thermal)'
        else:
            ticket = format_ticket_dotmatrix(
                no_transaksi, barang, barang_id, barang_lot,
                kustomer, kustomer_id, supplier, supplier_id,
                weight, keterangan, operator,
                company_name, company_address, company_phone, company_email
            )
            printer_type_display = 'Epson LX-310 (Dot Matrix)'
        
        # Kirim ke printer
        result = send_to_printer(printer_name, ticket, printer_type)
        
        if result['status'] == 'success':
            # Log aktivitas
            try:
                UserActivity.objects.create(
                    user=request.user,
                    action='Print Ticket',
                    details=f'Printed ticket: {no_transaksi} on {printer_type_display}',
                    ip_address=request.META.get('REMOTE_ADDR', '')
                )
            except:
                pass
            
            return JsonResponse({
                'status': 'success',
                'message': f'Ticket printed successfully on {printer_type_display}',
                'printer': printer_name,
                'printer_type': printer_type_display,
                'no_transaksi': no_transaksi
            })
        else:
            return JsonResponse({
                'status': 'error',
                'message': result.get('message', 'Print failed'),
                'printer': printer_name
            }, status=500)
            
    except json.JSONDecodeError as e:
        return JsonResponse({
            'status': 'error',
            'message': f'Invalid JSON: {str(e)}'
        }, status=400)
    except Exception as e:
        logger.error(f"Print ticket error: {e}", exc_info=True)
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=500)


# ============================================
# DETEKSI PRINTER (API)
# ============================================

@login_required
@csrf_exempt
def detect_printers(request):
    """API untuk mendeteksi printer yang tersedia"""
    try:
        printers = []
        
        # Coba melalui QZ Tray
        try:
            response = requests.get('http://localhost:8182/api/printer/list', timeout=3)
            if response.status_code == 200:
                qz_printers = response.json()
                for p in qz_printers:
                    printers.append({
                        'name': p.get('name', ''),
                        'is_default': p.get('isDefault', False),
                        'type': detect_printer_type(p.get('name', ''))
                    })
        except Exception as e:
            logger.warning(f"QZ Tray detect error: {e}")
        
        # Coba melalui Windows (jika di Windows dan win32print tersedia)
        if WIN32_AVAILABLE and os.name == 'nt' and not printers:
            try:
                if win32print is not None:
                    win_printers = win32print.EnumPrinters(
                        win32print.PRINTER_ENUM_LOCAL | win32print.PRINTER_ENUM_CONNECTIONS
                    )
                    for p in win_printers:
                        name = p[2]
                        printers.append({
                            'name': name,
                            'is_default': False,
                            'type': detect_printer_type(name)
                        })
            except Exception as e:
                logger.warning(f"Windows detect error: {e}")
        
        # Tandai printer default
        if printers:
            printers[0]['is_default'] = True
        
        return JsonResponse({
            'status': 'success',
            'printers': printers,
            'count': len(printers)
        })
        
    except Exception as e:
        logger.error(f"Detect printers error: {e}")
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=500)


# ============================================
# END OF PRINT FUNCTIONS
# ============================================

def ensure_default_weight_data():
    """Pastikan ada data weight = 0 di database"""
    try:
        # Cek apakah ada data terakhir
        latest = WeightData.objects.all().order_by('-created_at').first()
        if not latest:
            # Buat data default
            WeightData.objects.create(
                entry_id='0',
                weight=0,
                latitude=-6.152256,
                longitude=106.694091,
                created_at=timezone.now()
            )
            logger.info("Default weight data created")
    except Exception as e:
        logger.error(f"Error creating default weight: {e}") 
    
# ============================================
# REPORT UPDATE VIEW
# ============================================

@login_required
def report_update(request, pk):
    """Update report with FDA compliance"""
    if request.method != 'POST':
        return JsonResponse({
            'status': 'error',
            'message': 'Method not allowed'
        }, status=405)
    
    try:
        data = json.loads(request.body)
        password = data.get('password')
        
        if not password:
            return JsonResponse({
                'status': 'error',
                'message': 'Password required for FDA compliance'
            }, status=401)
        
        if not request.user.check_password(password):
            return JsonResponse({
                'status': 'error',
                'message': 'Invalid password'
            }, status=401)
        
        report = Report.objects.get(id=pk, report_type='Transaction')
        
        # Update data
        report_data = report.data if isinstance(report.data, dict) else {}
        
        # Update fields yang dikirim
        if 'no_transaksi' in data:
            report_data['no_transaksi'] = data.get('no_transaksi')
        if 'barang' in data:
            report_data['barang'] = data.get('barang')
        if 'barang_id' in data:
            report_data['barang_id'] = data.get('barang_id')
        if 'barang_lot' in data:
            report_data['barang_lot'] = data.get('barang_lot')
        if 'kustomer' in data:
            report_data['kustomer'] = data.get('kustomer')
        if 'kustomer_id' in data:
            report_data['kustomer_id'] = data.get('kustomer_id')
        if 'supplier' in data:
            report_data['supplier'] = data.get('supplier')
        if 'supplier_id' in data:
            report_data['supplier_id'] = data.get('supplier_id')
        if 'weight' in data:
            report_data['weight'] = str(data.get('weight', 0))
        if 'keterangan' in data:
            report_data['keterangan'] = data.get('keterangan')
        
        report.data = report_data
        report.save()
        
        try:
            UserActivity.objects.create(
                user=request.user,
                action='Update Report',
                details=f'Updated report ID: {pk}',
                ip_address=request.META.get('REMOTE_ADDR', '')
            )
        except:
            pass
        
        return JsonResponse({
            'status': 'success',
            'message': 'Report updated successfully'
        })
        
    except Report.DoesNotExist:
        return JsonResponse({
            'status': 'error',
            'message': 'Report not found'
        }, status=404)
    except Exception as e:
        logger.error(f"Update report error: {e}")
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=500)

# ============================================
# HELPER FUNCTIONS
# ============================================

def get_photo_url(file_path):
    """
    Helper function untuk mendapatkan URL foto dengan aman,
    baik dari FieldFile maupun string.
    """
    if not file_path:
        return ''
    
    try:
        if hasattr(file_path, 'url'):
            return file_path.url
        
        file_path_str = str(file_path)
        if not file_path_str:
            return ''
        
        if file_path_str.startswith('http'):
            return file_path_str
        
        file_path_str = file_path_str.lstrip('/')
        return settings.MEDIA_URL + file_path_str
        
    except Exception as e:
        logger.error(f"Error getting photo URL: {e}")
        return ''

def check_db_connection():
    """Cek koneksi database"""
    try:
        connection.ensure_connection()
        return True
    except Exception as e:
        logger.error(f"Database connection error: {e}")
        return False

# ============================================
# AUTHENTICATION VIEWS
# ============================================
@csrf_protect
def login_view(request):
    """Halaman login"""
    if request.user.is_authenticated:
        return redirect('dashboard:dashboard')
    
    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(username=username, password=password)
            if user is not None:
                login(request, user)
                try:
                    UserActivity.objects.create(
                        user=user,
                        action='Login',
                        details='User logged in successfully',
                        ip_address=request.META.get('REMOTE_ADDR')
                    )
                except:
                    pass
                return redirect('dashboard:dashboard')
    else:
        form = AuthenticationForm()
    
    company = CompanyProfile.objects.first()
    return render(request, 'dashboard/login.html', {
        'form': form,
        'company': company
    })

@login_required
def logout_view(request):
    """Logout user"""
    try:
        UserActivity.objects.create(
            user=request.user,
            action='Logout',
            details='User logged out',
            ip_address=request.META.get('REMOTE_ADDR')
        )
    except:
        pass
    logout(request)
    return redirect('dashboard:login')

# ============================================
# DASHBOARD VIEWS
# ============================================

@login_required
def dashboard(request):
    """Dashboard utama dengan data lengkap"""
    db_connected = check_db_connection()
    
    if db_connected:
        try:
            latest_data = WeightData.objects.all().order_by('-created_at')[:10]
            total_weights = WeightData.objects.count()
            avg_weight = WeightData.objects.aggregate(avg=Avg('weight'))['avg']
            max_weight = WeightData.objects.aggregate(max=Max('weight'))['max']
            min_weight = WeightData.objects.aggregate(min=Min('weight'))['min']
            
            today = timezone.now().date()
            today_data = WeightData.objects.filter(created_at__date=today)
            today_data_count = today_data.count()
            weight_history = WeightData.objects.all().order_by('-created_at')[:50]
        except Exception as e:
            logger.error(f"Dashboard data error: {e}")
            latest_data = []
            total_weights = 0
            avg_weight = 0
            max_weight = 0
            min_weight = 0
            today_data_count = 0
            weight_history = []
    else:
        latest_data = []
        total_weights = 0
        avg_weight = 0
        max_weight = 0
        min_weight = 0
        today_data_count = 0
        weight_history = []
    
    try:
        today_history = Report.objects.filter(
            report_type='Transaction',
            created_at__date=timezone.now().date()
        ).order_by('-created_at')[:50]
    except:
        today_history = []
    
    company = CompanyProfile.objects.first()
    barang_list = Barang.objects.all().order_by('nama_barang') if db_connected else []
    kustomer_list = Kustomer.objects.all().order_by('nama_kustomer') if db_connected else []
    supplier_list = Supplier.objects.all().order_by('nama_supplier') if db_connected else []
    
    context = {
        'latest_data': latest_data,
        'total_weights': total_weights or 0,
        'avg_weight': avg_weight or 0,
        'max_weight': max_weight or 0,
        'min_weight': min_weight or 0,
        'today_data_count': today_data_count,
        'today_history': today_history,
        'company': company,
        'weight_history': weight_history,
        'barang_list': barang_list,
        'kustomer_list': kustomer_list,
        'supplier_list': supplier_list,
        'db_connected': db_connected,
    }
    return render(request, 'dashboard/dashboard.html', context)

# ============================================
# API VIEWS - REALTIME
# ============================================

@login_required
@csrf_exempt
def get_weight_history(request):
    """Get weight history data"""
    try:
        if not check_db_connection():
            api_setting = APISetting.objects.first()
            if api_setting:
                try:
                    url = f"https://api.thingspeak.com/channels/{api_setting.channel_id}/feeds.json"
                    params = {'api_key': api_setting.read_api_key, 'results': 30}
                    response = requests.get(url, params=params, timeout=10)
                    if response.status_code == 200:
                        feeds = response.json().get('feeds', [])
                        data = []
                        for feed in feeds[::-1]:
                            data.append({
                                'timestamp': feed.get('created_at', ''),
                                'weight': float(feed.get('field1', 0))
                            })
                        return JsonResponse({'status': 'success', 'data': data, 'source': 'thingspeak'})
                except:
                    pass
            return JsonResponse({'status': 'success', 'data': [], 'source': 'empty'})
        
        history = WeightData.objects.all().order_by('-created_at')[:30]
        data = []
        for item in history:
            data.append({
                'timestamp': item.created_at.strftime('%H:%M:%S'),
                'weight': float(item.weight)
            })
        return JsonResponse({'status': 'success', 'data': data, 'source': 'database'})
        
    except Exception as e:
        logger.error(f"Get history error: {e}")
        return JsonResponse({'status': 'error', 'message': str(e), 'data': []})

@login_required
@csrf_exempt
def get_thingspeak_data(request):
    """Get data langsung dari ThingSpeak"""
    try:
        api_setting = APISetting.objects.first()
        if not api_setting:
            return JsonResponse({'status': 'error', 'message': 'API setting not found'})
        
        url = f"https://api.thingspeak.com/channels/{api_setting.channel_id}/feeds.json"
        params = {
            'api_key': api_setting.read_api_key,
            'results': 5
        }
        response = requests.get(url, params=params, timeout=10)
        
        if response.status_code == 200:
            data = response.json()
            feeds = data.get('feeds', [])
            
            if feeds:
                latest = feeds[0]
                result = {
                    'entry_id': latest.get('entry_id'),
                    'weight': float(latest.get('field1', 0)),
                    'latitude': float(latest.get('field2', 0)),
                    'longitude': float(latest.get('field3', 0)),
                    'created_at': latest.get('created_at'),
                    'has_object': float(latest.get('field1', 0)) > 0
                }
                return JsonResponse({'status': 'success', 'data': result})
        
        return JsonResponse({'status': 'error', 'message': 'Failed to fetch from ThingSpeak'})
    except requests.exceptions.Timeout:
        return JsonResponse({'status': 'error', 'message': 'Connection timeout'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})

@login_required
@csrf_exempt
def get_latest_weight(request):
    """Get latest weight data - FILTER DATA BASI"""
    try:
        now = timezone.now()
        
        if not check_db_connection():
            # Coba dari ThingSpeak langsung
            api_setting = APISetting.objects.first()
            if api_setting:
                try:
                    url = f"https://api.thingspeak.com/channels/{api_setting.channel_id}/feeds.json"
                    params = {'api_key': api_setting.read_api_key, 'results': 2}
                    response = requests.get(url, params=params, timeout=10)
                    if response.status_code == 200:
                        feeds = response.json().get('feeds', [])
                        if feeds:
                            latest = feeds[0]
                            weight = float(latest.get('field1', 0))
                            created_at = latest.get('created_at')
                            
                            # CEK WAKTU DATA
                            if created_at:
                                try:
                                    data_time = datetime.strptime(created_at, '%Y-%m-%dT%H:%M:%SZ')
                                    data_time = timezone.make_aware(data_time)
                                    diff = (now - data_time).total_seconds()
                                    
                                    # JIKA DATA LEBIH DARI 30 DETIK, TAMPILKAN 0
                                    if diff > 30:
                                        weight = 0
                                except:
                                    pass
                            
                            data = {
                                'entry_id': int(latest.get('entry_id', 0)) if weight > 0 else None,
                                'weight': weight,
                                'latitude': float(latest.get('field2', -6.152256)),
                                'longitude': float(latest.get('field3', 106.694091)),
                                'created_at': latest.get('created_at'),
                                'has_object': weight > 0,
                                'source': 'thingspeak'
                            }
                            return JsonResponse({'status': 'success', 'data': data})
                except:
                    pass
            
            # Data default
            data = {
                'entry_id': None,
                'weight': 0.00,
                'latitude': -6.152256,
                'longitude': 106.694091,
                'created_at': None,
                'has_object': False,
                'source': 'default'
            }
            return JsonResponse({'status': 'success', 'data': data})
        
        # PRIORITAS: Ambil dari database dulu
        latest = WeightData.objects.all().order_by('-entry_id').first()
        
        if latest:
            weight = float(latest.weight)
            
            # CEK WAKTU DATA DARI DATABASE
            if latest.created_at:
                diff = (now - latest.created_at).total_seconds()
                
                # JIKA DATA LEBIH DARI 30 DETIK, COBA AMBIL DARI THINGSPEAK
                if diff > 30:
                    api_setting = APISetting.objects.first()
                    if api_setting and api_setting.channel_id and api_setting.read_api_key:
                        try:
                            url = f"https://api.thingspeak.com/channels/{api_setting.channel_id}/feeds/last.json"
                            params = {'api_key': api_setting.read_api_key}
                            response = requests.get(url, params=params, timeout=5)
                            if response.status_code == 200:
                                feed = response.json()
                                if feed:
                                    ts_weight = float(feed.get('field1', 0))
                                    ts_created = feed.get('created_at')
                                    
                                    if ts_created:
                                        try:
                                            ts_time = datetime.strptime(ts_created, '%Y-%m-%dT%H:%M:%SZ')
                                            ts_time = timezone.make_aware(ts_time)
                                            ts_diff = (now - ts_time).total_seconds()
                                            
                                            # JIKA DATA THINGSPEAK LEBIH BARU, PAKAI ITU
                                            if ts_diff < 30 and ts_diff < diff:
                                                weight = ts_weight
                                                # Update database
                                                WeightData.objects.update_or_create(
                                                    entry_id=str(feed.get('entry_id')) if feed.get('entry_id') else '0',
                                                    defaults={
                                                        'weight': weight,
                                                        'latitude': float(feed.get('field2', -6.152256)),
                                                        'longitude': float(feed.get('field3', 106.694091)),
                                                        'created_at': ts_time
                                                    }
                                                )
                                        except:
                                            pass
                        except:
                            pass
            
            data = {
                'entry_id': latest.entry_id if weight > 0 else None,
                'weight': weight,
                'latitude': float(latest.latitude),
                'longitude': float(latest.longitude),
                'created_at': latest.created_at.isoformat() if latest.created_at else None,
                'has_object': weight > 0,
                'source': 'database'
            }
        else:
            data = {
                'entry_id': None,
                'weight': 0.00,
                'latitude': -6.152256,
                'longitude': 106.694091,
                'created_at': None,
                'has_object': False,
                'source': 'empty'
            }
            
        return JsonResponse({'status': 'success', 'data': data})
        
    except Exception as e:
        logger.error(f"Get latest weight error: {e}")
        return JsonResponse({
            'status': 'error', 
            'message': str(e),
            'data': {
                'entry_id': None,
                'weight': 0.00,
                'latitude': -6.152256,
                'longitude': 106.694091,
                'created_at': None,
                'has_object': False
            }
        })

@login_required
@csrf_exempt
def sync_weight_data(request):
    """Sync data from ThingSpeak - SIMPAN JUGA WEIGHT=0"""
    try:
        # CEK CACHE
        cache_key = f'sync_data_{request.user.id}'
        cached_result = cache.get(cache_key)
        
        if cached_result:
            return JsonResponse(cached_result)
        
        if not check_db_connection():
            return JsonResponse({
                'status': 'error',
                'message': 'Database connection failed.'
            }, status=500)
        
        api_setting = APISetting.objects.first()
        if not api_setting:
            return JsonResponse({
                'status': 'error',
                'message': 'API setting not found.'
            }, status=404)
        
        if not api_setting.channel_id or not api_setting.read_api_key:
            return JsonResponse({
                'status': 'error',
                'message': 'Channel ID atau Read API Key kosong.'
            }, status=400)
        
        # AMBIL HANYA 1 DATA TERBARU
        url = f"https://api.thingspeak.com/channels/{api_setting.channel_id}/feeds/last.json"
        params = {'api_key': api_setting.read_api_key}
        
        response = requests.get(url, params=params, timeout=3)
        
        if response.status_code == 200:
            feed = response.json()
            
            if feed:
                entry_id = feed.get('entry_id')
                weight = float(feed.get('field1', 0))
                latitude = float(feed.get('field2', -6.152256))
                longitude = float(feed.get('field3', 106.694091))
                created_at_str = feed.get('created_at')
                
                if created_at_str:
                    try:
                        created_at = datetime.strptime(created_at_str, '%Y-%m-%dT%H:%M:%SZ')
                        created_at = timezone.make_aware(created_at)
                    except:
                        created_at = timezone.now()
                else:
                    created_at = timezone.now()
                
                # ✅ SIMPAN KE DATABASE (TERMASUK WEIGHT = 0)
                obj, created = WeightData.objects.update_or_create(
                    entry_id=str(entry_id) if entry_id else '0',
                    defaults={
                        'weight': weight,
                        'latitude': latitude,
                        'longitude': longitude,
                        'created_at': created_at
                    }
                )
                
                result = {
                    'status': 'success',
                    'message': f'Sync successful. Weight: {weight} kg',
                    'latest_weight': weight,
                    'latest_entry_id': str(entry_id) if entry_id else '--',
                    'created': created
                }
                
                # SIMPAN CACHE (5 detik)
                cache.set(cache_key, result, 5)
                
                return JsonResponse(result)
            else:
                return JsonResponse({
                    'status': 'warning',
                    'message': 'No data from ThingSpeak channel'
                })
        else:
            return JsonResponse({
                'status': 'error',
                'message': f'ThingSpeak API error: HTTP {response.status_code}'
            })
            
    except requests.exceptions.Timeout:
        return JsonResponse({
            'status': 'error',
            'message': 'Connection timeout.'
        }, status=408)
    except requests.exceptions.ConnectionError:
        return JsonResponse({
            'status': 'error',
            'message': 'Cannot connect to ThingSpeak.'
        }, status=503)
    except Exception as e:
        logger.error(f"API sync error: {e}", exc_info=True)
        return JsonResponse({
            'status': 'error',
            'message': f'Sync error: {str(e)}'
        }, status=500)

@login_required
@csrf_exempt
def test_api_connection(request):
    """Test ThingSpeak API connection"""
    try:
        api_setting = APISetting.objects.first()
        if not api_setting:
            return JsonResponse({
                'status': 'error', 
                'message': 'API setting not found'
            }, status=404)
        
        if not api_setting.channel_id or not api_setting.read_api_key:
            return JsonResponse({
                'status': 'error',
                'message': 'Channel ID atau Read API Key kosong. Silakan lengkapi di pengaturan.'
            }, status=400)
        
        url = f"https://api.thingspeak.com/channels/{api_setting.channel_id}/feeds.json"
        params = {
            'api_key': api_setting.read_api_key,
            'results': 1
        }
        
        response = requests.get(url, params=params, timeout=10)
        
        if response.status_code == 200:
            data = response.json()
            if data.get('feeds'):
                try:
                    api_setting.last_test = timezone.now()
                    api_setting.save()
                    UserActivity.objects.create(
                        user=request.user,
                        action='Test API',
                        details='API connection test successful',
                        ip_address=request.META.get('REMOTE_ADDR')
                    )
                except:
                    pass
                
                return JsonResponse({
                    'status': 'success',
                    'message': 'Connection successful!',
                    'data': data,
                    'channel_id': api_setting.channel_id,
                    'total_feeds': len(data.get('feeds', []))
                })
            else:
                return JsonResponse({
                    'status': 'warning',
                    'message': 'Connection successful but no data in channel',
                    'channel_id': api_setting.channel_id
                })
        else:
            return JsonResponse({
                'status': 'error',
                'message': f'Connection failed. Status code: {response.status_code}'
            }, status=response.status_code)
            
    except requests.exceptions.Timeout:
        return JsonResponse({
            'status': 'error', 
            'message': 'Connection timeout'
        }, status=408)
    except requests.exceptions.ConnectionError:
        return JsonResponse({
            'status': 'error', 
            'message': 'Cannot connect to ThingSpeak'
        }, status=503)
    except Exception as e:
        logger.error(f"Test API error: {e}")
        return JsonResponse({
            'status': 'error', 
            'message': str(e)
        }, status=500)

# ============================================
# DATA MANAGEMENT VIEWS
# ============================================

@login_required
def data_barang(request):
    """Manage Barang data with search functionality"""
    try:
        # Get search parameters
        search_nama = request.GET.get('search_nama', '').strip()
        search_kategori = request.GET.get('search_kategori', '').strip()
        search_id = request.GET.get('search_id', '').strip()
        
        barang_list = Barang.objects.all().order_by('-created_at')
        
        # Apply filters
        if search_nama:
            barang_list = barang_list.filter(nama_barang__icontains=search_nama)
        if search_kategori:
            barang_list = barang_list.filter(kategori__icontains=search_kategori)
        if search_id:
            barang_list = barang_list.filter(id_barang__icontains=search_id)
        
        paginator = Paginator(barang_list, 20)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)
    except Exception as e:
        logger.error(f"Data barang error: {e}")
        page_obj = []
    
    if request.method == 'POST':
        form = BarangForm(request.POST)
        if form.is_valid():
            barang = form.save()
            try:
                UserActivity.objects.create(
                    user=request.user,
                    action='Add Barang',
                    details=f'Added barang: {barang.nama_barang} (ID: {barang.id_barang})',
                    ip_address=request.META.get('REMOTE_ADDR')
                )
            except:
                pass
            return redirect('dashboard:data_barang')
    else:
        form = BarangForm()
    
    return render(request, 'dashboard/data_barang.html', {
        'page_obj': page_obj,
        'form': form,
    })

@login_required
def export_barang_excel(request):
    """Export barang data to Excel with FDA Part 11 compliance"""
    try:
        password = request.GET.get('password')
        if not password:
            return JsonResponse({
                'status': 'error',
                'message': 'Password required for FDA Part 11 compliance'
            }, status=401)
        
        if not request.user.check_password(password):
            return JsonResponse({
                'status': 'error',
                'message': 'Invalid password'
            }, status=401)
        
        search_nama = request.GET.get('search_nama', '').strip()
        search_kategori = request.GET.get('search_kategori', '').strip()
        search_id = request.GET.get('search_id', '').strip()
        
        barang_list = Barang.objects.all().order_by('-created_at')
        
        if search_nama:
            barang_list = barang_list.filter(nama_barang__icontains=search_nama)
        if search_kategori:
            barang_list = barang_list.filter(kategori__icontains=search_kategori)
        if search_id:
            barang_list = barang_list.filter(id_barang__icontains=search_id)
        
        export_data = []
        for idx, item in enumerate(barang_list, 1):
            export_data.append({
                'No': idx,
                'ID Barang': item.id_barang,
                'Nama Barang': item.nama_barang,
                'Lot': item.lot,
                'Kategori': item.kategori or '-',
                'Deskripsi': item.deskripsi or '-',
                'Tanggal Dibuat': item.created_at.strftime('%Y-%m-%d %H:%M:%S') if item.created_at else '',
            })
        
        df = pd.DataFrame(export_data)
        
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Data Barang', index=False)
            
            workbook = writer.book
            worksheet = writer.sheets['Data Barang']
            
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            header_font = Font(bold=True, size=11, color='FFFFFF')
            header_fill = PatternFill(start_color='0D47A1', end_color='0D47A1', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center')
            
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
        
        try:
            UserActivity.objects.create(
                user=request.user,
                action='Export Barang',
                details=f'Exported {len(barang_list)} barang records to Excel',
                ip_address=request.META.get('REMOTE_ADDR', '')
            )
        except:
            pass
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'data_barang_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
        
    except Exception as e:
        logger.error(f"Export barang error: {e}", exc_info=True)
        return JsonResponse({
            'status': 'error',
            'message': f'Export failed: {str(e)}'
        }, status=500)

@login_required
def import_barang_excel(request):
    """Import barang data from Excel with FDA Part 11 compliance"""
    try:
        password = request.POST.get('password')
        if not password:
            return JsonResponse({
                'status': 'error',
                'message': 'Password required for FDA Part 11 compliance'
            }, status=401)
        
        if not request.user.check_password(password):
            return JsonResponse({
                'status': 'error',
                'message': 'Invalid password'
            }, status=401)
        
        if 'file' not in request.FILES:
            return JsonResponse({
                'status': 'error',
                'message': 'No file uploaded'
            }, status=400)
        
        file = request.FILES['file']
        
        # Read Excel file
        try:
            df = pd.read_excel(file)
        except Exception as e:
            return JsonResponse({
                'status': 'error',
                'message': f'Failed to read file: {str(e)}'
            }, status=400)
        
        required_columns = ['ID Barang', 'Nama Barang', 'Lot']
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            return JsonResponse({
                'status': 'error',
                'message': f'Column(s) not found: {", ".join(missing_columns)}'
            }, status=400)
        
        imported_count = 0
        updated_count = 0
        error_messages = []
        
        for index, row in df.iterrows():
            try:
                id_barang = str(row.get('ID Barang', '')).strip()
                nama_barang = str(row.get('Nama Barang', '')).strip()
                lot = str(row.get('Lot', '')).strip()
                
                # Handle NaN values
                kategori = ''
                if pd.notna(row.get('Kategori')):
                    kategori = str(row.get('Kategori')).strip()
                
                deskripsi = ''
                if pd.notna(row.get('Deskripsi')):
                    deskripsi = str(row.get('Deskripsi')).strip()
                
                # FIX: Deklarasikan row_num di sini
                row_num = int(index) + 2 if isinstance(index, (int, float)) else '?'
                
                if not id_barang or not nama_barang or not lot:
                    error_messages.append(f"Row {row_num}: ID Barang, Nama Barang, and Lot are required")
                    continue
                
                obj, created = Barang.objects.update_or_create(
                    id_barang=id_barang,
                    defaults={
                        'nama_barang': nama_barang,
                        'lot': lot,
                        'kategori': kategori,
                        'deskripsi': deskripsi
                    }
                )
                
                if created:
                    imported_count += 1
                else:
                    updated_count += 1
                    
            except Exception as e:
                # FIX: Gunakan index yang aman
                row_num = int(index) + 2 if isinstance(index, (int, float)) else '?'
                error_messages.append(f"Row {row_num}: {str(e)}")
        
        try:
            UserActivity.objects.create(
                user=request.user,
                action='Import Barang',
                details=f'Imported {imported_count} new, updated {updated_count} barang records',
                ip_address=request.META.get('REMOTE_ADDR', '')
            )
        except:
            pass
        
        message_parts = []
        if imported_count > 0:
            message_parts.append(f'{imported_count} data baru ditambahkan')
        if updated_count > 0:
            message_parts.append(f'{updated_count} data diperbarui')
        if error_messages:
            message_parts.append(f'{len(error_messages)} error')
        
        message = ' | '.join(message_parts) if message_parts else 'Tidak ada data yang diimport'
        
        return JsonResponse({
            'status': 'success',
            'message': message,
            'imported': imported_count,
            'updated': updated_count,
            'errors': error_messages[:10] if error_messages else []
        })
        
    except Exception as e:
        logger.error(f"Import barang error: {e}", exc_info=True)
        return JsonResponse({
            'status': 'error',
            'message': f'Import failed: {str(e)}'
        }, status=500)

# ============================================
# EXPORT KUSTOMER EXCEL
# ============================================

@login_required
def export_kustomer_excel(request):
    """Export kustomer data to Excel with FDA Part 11 compliance"""
    try:
        password = request.GET.get('password')
        if not password:
            return JsonResponse({
                'status': 'error',
                'message': 'Password required for FDA Part 11 compliance'
            }, status=401)
        
        if not request.user.check_password(password):
            return JsonResponse({
                'status': 'error',
                'message': 'Invalid password'
            }, status=401)
        
        search_nama = request.GET.get('search_nama', '').strip()
        search_id = request.GET.get('search_id', '').strip()
        
        kustomer_list = Kustomer.objects.all().order_by('-created_at')
        
        if search_nama:
            kustomer_list = kustomer_list.filter(nama_kustomer__icontains=search_nama)
        if search_id:
            kustomer_list = kustomer_list.filter(id_kustomer__icontains=search_id)
        
        export_data = []
        for idx, item in enumerate(kustomer_list, 1):
            export_data.append({
                'No': idx,
                'ID Kustomer': item.id_kustomer,
                'Nama Kustomer': item.nama_kustomer,
                'Alamat': item.alamat or '-',
                'Telepon': item.telepon or '-',
                'Email': item.email or '-',
                'Tanggal Dibuat': item.created_at.strftime('%Y-%m-%d %H:%M:%S') if item.created_at else '',
            })
        
        df = pd.DataFrame(export_data)
        
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Data Kustomer', index=False)
            
            workbook = writer.book
            worksheet = writer.sheets['Data Kustomer']
            
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            header_font = Font(bold=True, size=11, color='FFFFFF')
            header_fill = PatternFill(start_color='1B7A34', end_color='1B7A34', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center')
            
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
        
        try:
            UserActivity.objects.create(
                user=request.user,
                action='Export Kustomer',
                details=f'Exported {len(kustomer_list)} kustomer records to Excel',
                ip_address=request.META.get('REMOTE_ADDR', '')
            )
        except:
            pass
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'data_kustomer_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
        
    except Exception as e:
        logger.error(f"Export kustomer error: {e}", exc_info=True)
        return JsonResponse({
            'status': 'error',
            'message': f'Export failed: {str(e)}'
        }, status=500)

# ============================================
# IMPORT KUSTOMER EXCEL
# ============================================

@login_required
def import_kustomer_excel(request):
    """Import kustomer data from Excel with FDA Part 11 compliance"""
    try:
        password = request.POST.get('password')
        if not password:
            return JsonResponse({
                'status': 'error',
                'message': 'Password required for FDA Part 11 compliance'
            }, status=401)
        
        if not request.user.check_password(password):
            return JsonResponse({
                'status': 'error',
                'message': 'Invalid password'
            }, status=401)
        
        if 'file' not in request.FILES:
            return JsonResponse({
                'status': 'error',
                'message': 'No file uploaded'
            }, status=400)
        
        file = request.FILES['file']
        
        try:
            df = pd.read_excel(file)
        except Exception as e:
            return JsonResponse({
                'status': 'error',
                'message': f'Failed to read file: {str(e)}'
            }, status=400)
        
        required_columns = ['ID Kustomer', 'Nama Kustomer']
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            return JsonResponse({
                'status': 'error',
                'message': f'Column(s) not found: {", ".join(missing_columns)}'
            }, status=400)
        
        imported_count = 0
        updated_count = 0
        error_messages = []
        
        for index, row in df.iterrows():
            try:
                id_kustomer = str(row.get('ID Kustomer', '')).strip()
                nama_kustomer = str(row.get('Nama Kustomer', '')).strip()
                alamat = str(row.get('Alamat', '')).strip() if pd.notna(row.get('Alamat')) else ''
                telepon = str(row.get('Telepon', '')).strip() if pd.notna(row.get('Telepon')) else ''
                email = str(row.get('Email', '')).strip() if pd.notna(row.get('Email')) else ''
                
                if not id_kustomer or not nama_kustomer:
                    row_num = int(index) + 2 if isinstance(index, (int, float)) else '?'
                    error_messages.append(f"Row {row_num}: ID Kustomer and Nama Kustomer are required")
                    continue
                
                obj, created = Kustomer.objects.update_or_create(
                    id_kustomer=id_kustomer,
                    defaults={
                        'nama_kustomer': nama_kustomer,
                        'alamat': alamat,
                        'telepon': telepon,
                        'email': email
                    }
                )
                
                if created:
                    imported_count += 1
                else:
                    updated_count += 1
                    
            except Exception as e:
                row_num = int(index) + 2 if isinstance(index, (int, float)) else '?'
                error_messages.append(f"Row {row_num}: {str(e)}")
        
        try:
            UserActivity.objects.create(
                user=request.user,
                action='Import Kustomer',
                details=f'Imported {imported_count} new, updated {updated_count} kustomer records',
                ip_address=request.META.get('REMOTE_ADDR', '')
            )
        except:
            pass
        
        message_parts = []
        if imported_count > 0:
            message_parts.append(f'{imported_count} data baru ditambahkan')
        if updated_count > 0:
            message_parts.append(f'{updated_count} data diperbarui')
        if error_messages:
            message_parts.append(f'{len(error_messages)} error')
        
        message = ' | '.join(message_parts) if message_parts else 'Tidak ada data yang diimport'
        
        return JsonResponse({
            'status': 'success',
            'message': message,
            'imported': imported_count,
            'updated': updated_count,
            'errors': error_messages[:10] if error_messages else []
        })
        
    except Exception as e:
        logger.error(f"Import kustomer error: {e}", exc_info=True)
        return JsonResponse({
            'status': 'error',
            'message': f'Import failed: {str(e)}'
        }, status=500)

# ============================================
# EXPORT SUPPLIER EXCEL
# ============================================

@login_required
def export_supplier_excel(request):
    """Export supplier data to Excel with FDA Part 11 compliance"""
    try:
        password = request.GET.get('password')
        if not password:
            return JsonResponse({
                'status': 'error',
                'message': 'Password required for FDA Part 11 compliance'
            }, status=401)
        
        if not request.user.check_password(password):
            return JsonResponse({
                'status': 'error',
                'message': 'Invalid password'
            }, status=401)
        
        search_nama = request.GET.get('search_nama', '').strip()
        search_id = request.GET.get('search_id', '').strip()
        
        supplier_list = Supplier.objects.all().order_by('-created_at')
        
        if search_nama:
            supplier_list = supplier_list.filter(nama_supplier__icontains=search_nama)
        if search_id:
            supplier_list = supplier_list.filter(id_supplier__icontains=search_id)
        
        export_data = []
        for idx, item in enumerate(supplier_list, 1):
            export_data.append({
                'No': idx,
                'ID Supplier': item.id_supplier,
                'Nama Supplier': item.nama_supplier,
                'Alamat': item.alamat or '-',
                'Telepon': item.telepon or '-',
                'Email': item.email or '-',
                'Tanggal Dibuat': item.created_at.strftime('%Y-%m-%d %H:%M:%S') if item.created_at else '',
            })
        
        df = pd.DataFrame(export_data)
        
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Data Supplier', index=False)
            
            workbook = writer.book
            worksheet = writer.sheets['Data Supplier']
            
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            header_font = Font(bold=True, size=11, color='FFFFFF')
            header_fill = PatternFill(start_color='0D47A1', end_color='0D47A1', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center')
            
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
        
        try:
            UserActivity.objects.create(
                user=request.user,
                action='Export Supplier',
                details=f'Exported {len(supplier_list)} supplier records to Excel',
                ip_address=request.META.get('REMOTE_ADDR', '')
            )
        except:
            pass
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'data_supplier_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
        
    except Exception as e:
        logger.error(f"Export supplier error: {e}", exc_info=True)
        return JsonResponse({
            'status': 'error',
            'message': f'Export failed: {str(e)}'
        }, status=500)

# ============================================
# IMPORT SUPPLIER EXCEL
# ============================================

@login_required
def import_supplier_excel(request):
    """Import supplier data from Excel with FDA Part 11 compliance"""
    try:
        password = request.POST.get('password')
        if not password:
            return JsonResponse({
                'status': 'error',
                'message': 'Password required for FDA Part 11 compliance'
            }, status=401)
        
        if not request.user.check_password(password):
            return JsonResponse({
                'status': 'error',
                'message': 'Invalid password'
            }, status=401)
        
        if 'file' not in request.FILES:
            return JsonResponse({
                'status': 'error',
                'message': 'No file uploaded'
            }, status=400)
        
        file = request.FILES['file']
        
        try:
            df = pd.read_excel(file)
        except Exception as e:
            return JsonResponse({
                'status': 'error',
                'message': f'Failed to read file: {str(e)}'
            }, status=400)
        
        required_columns = ['ID Supplier', 'Nama Supplier']
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            return JsonResponse({
                'status': 'error',
                'message': f'Column(s) not found: {", ".join(missing_columns)}'
            }, status=400)
        
        imported_count = 0
        updated_count = 0
        error_messages = []
        
        for index, row in df.iterrows():
            try:
                id_supplier = str(row.get('ID Supplier', '')).strip()
                nama_supplier = str(row.get('Nama Supplier', '')).strip()
                alamat = str(row.get('Alamat', '')).strip() if pd.notna(row.get('Alamat')) else ''
                telepon = str(row.get('Telepon', '')).strip() if pd.notna(row.get('Telepon')) else ''
                email = str(row.get('Email', '')).strip() if pd.notna(row.get('Email')) else ''
                
                if not id_supplier or not nama_supplier:
                    row_num = int(index) + 2 if isinstance(index, (int, float)) else '?'
                    error_messages.append(f"Row {row_num}: ID Supplier and Nama Supplier are required")
                    continue
                
                obj, created = Supplier.objects.update_or_create(
                    id_supplier=id_supplier,
                    defaults={
                        'nama_supplier': nama_supplier,
                        'alamat': alamat,
                        'telepon': telepon,
                        'email': email
                    }
                )
                
                if created:
                    imported_count += 1
                else:
                    updated_count += 1
                    
            except Exception as e:
                row_num = int(index) + 2 if isinstance(index, (int, float)) else '?'
                error_messages.append(f"Row {row_num}: {str(e)}")
        
        try:
            UserActivity.objects.create(
                user=request.user,
                action='Import Supplier',
                details=f'Imported {imported_count} new, updated {updated_count} supplier records',
                ip_address=request.META.get('REMOTE_ADDR', '')
            )
        except:
            pass
        
        message_parts = []
        if imported_count > 0:
            message_parts.append(f'{imported_count} data baru ditambahkan')
        if updated_count > 0:
            message_parts.append(f'{updated_count} data diperbarui')
        if error_messages:
            message_parts.append(f'{len(error_messages)} error')
        
        message = ' | '.join(message_parts) if message_parts else 'Tidak ada data yang diimport'
        
        return JsonResponse({
            'status': 'success',
            'message': message,
            'imported': imported_count,
            'updated': updated_count,
            'errors': error_messages[:10] if error_messages else []
        })
        
    except Exception as e:
        logger.error(f"Import supplier error: {e}", exc_info=True)
        return JsonResponse({
            'status': 'error',
            'message': f'Import failed: {str(e)}'
        }, status=500)

# ============================================
# DATA KUSTOMER
# ============================================

@login_required
def data_kustomer(request):
    """Manage Kustomer data"""
    try:
        # Get search parameters
        search_nama = request.GET.get('search_nama', '').strip()
        search_alamat = request.GET.get('search_alamat', '').strip()
        search_id = request.GET.get('search_id', '').strip()
        
        kustomer_list = Kustomer.objects.all().order_by('-created_at')
        
        if search_nama:
            kustomer_list = kustomer_list.filter(nama_kustomer__icontains=search_nama)
        if search_alamat:
            kustomer_list = kustomer_list.filter(alamat__icontains=search_alamat)
        if search_id:
            kustomer_list = kustomer_list.filter(id_kustomer__icontains=search_id)
        
        paginator = Paginator(kustomer_list, 20)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)
    except:
        page_obj = []
    
    if request.method == 'POST':
        form = KustomerForm(request.POST)
        if form.is_valid():
            kustomer = form.save()
            try:
                UserActivity.objects.create(
                    user=request.user,
                    action='Add Kustomer',
                    details=f'Added kustomer: {kustomer.nama_kustomer} (ID: {kustomer.id_kustomer})',
                    ip_address=request.META.get('REMOTE_ADDR')
                )
            except:
                pass
            return redirect('dashboard:data_kustomer')
    else:
        form = KustomerForm()
    
    return render(request, 'dashboard/data_kustomer.html', {
        'page_obj': page_obj,
        'form': form,
    })

# ============================================
# DATA SUPPLIER
# ============================================

@login_required
def data_supplier(request):
    """Manage Supplier data"""
    try:
        # Get search parameters
        search_nama = request.GET.get('search_nama', '').strip()
        search_alamat = request.GET.get('search_alamat', '').strip()
        search_id = request.GET.get('search_id', '').strip()
        
        supplier_list = Supplier.objects.all().order_by('-created_at')
        
        if search_nama:
            supplier_list = supplier_list.filter(nama_supplier__icontains=search_nama)
        if search_alamat:
            supplier_list = supplier_list.filter(alamat__icontains=search_alamat)
        if search_id:
            supplier_list = supplier_list.filter(id_supplier__icontains=search_id)
        
        paginator = Paginator(supplier_list, 20)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)
    except:
        page_obj = []
    
    if request.method == 'POST':
        form = SupplierForm(request.POST)
        if form.is_valid():
            supplier = form.save()
            try:
                UserActivity.objects.create(
                    user=request.user,
                    action='Add Supplier',
                    details=f'Added supplier: {supplier.nama_supplier} (ID: {supplier.id_supplier})',
                    ip_address=request.META.get('REMOTE_ADDR')
                )
            except:
                pass
            return redirect('dashboard:data_supplier')
    else:
        form = SupplierForm()
    
    return render(request, 'dashboard/data_supplier.html', {
        'page_obj': page_obj,
        'form': form,
    })

# ============================================
# SETTINGS VIEWS
# ============================================

@login_required
def setting(request):
    """Settings page for API and Company"""
    api_setting = APISetting.objects.first()
    company = CompanyProfile.objects.first()
    
    api_form = APISettingForm(instance=api_setting)
    company_form = CompanyProfileForm(instance=company)
    
    if request.method == 'POST':
        if 'api_form' in request.POST:
            api_form = APISettingForm(request.POST, instance=api_setting)
            if api_form.is_valid():
                api_setting = api_form.save()
                try:
                    UserActivity.objects.create(
                        user=request.user,
                        action='Update API Settings',
                        details=f'Updated API settings for channel: {api_setting.channel_id}',
                        ip_address=request.META.get('REMOTE_ADDR')
                    )
                except:
                    pass
                return redirect('dashboard:setting')
        elif 'company_form' in request.POST:
            company_form = CompanyProfileForm(request.POST, request.FILES, instance=company)
            if company_form.is_valid():
                company = company_form.save()
                try:
                    UserActivity.objects.create(
                        user=request.user,
                        action='Update Company Profile',
                        details=f'Updated company profile: {company.name}',
                        ip_address=request.META.get('REMOTE_ADDR')
                    )
                except:
                    pass
                return redirect('dashboard:setting')
    
    return render(request, 'dashboard/setting.html', {
        'api_form': api_form,
        'company_form': company_form,
        'api_setting': api_setting,
        'company': company
    })
    
# ============================================
# VERIFY PASSWORD (FDA Part 11)
# ============================================

@login_required
def verify_password(request):
    """Verify user password for FDA compliance"""
    if request.method != 'POST':
        return JsonResponse({
            'status': 'error', 
            'message': 'Method not allowed'
        }, status=405)
    
    try:
        data = json.loads(request.body)
        password = data.get('password')
        
        if not password:
            return JsonResponse({
                'status': 'error', 
                'message': 'Password is required'
            }, status=400)
        
        if request.user.check_password(password):
            try:
                UserActivity.objects.create(
                    user=request.user,
                    action='Verify Password',
                    details='Password verified successfully',
                    ip_address=request.META.get('REMOTE_ADDR')
                )
            except:
                pass
            return JsonResponse({'status': 'success', 'message': 'Password verified'})
        else:
            return JsonResponse({
                'status': 'error', 
                'message': 'Invalid password'
            }, status=401)
            
    except json.JSONDecodeError:
        return JsonResponse({
            'status': 'error', 
            'message': 'Invalid JSON data'
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'status': 'error', 
            'message': str(e)
        }, status=500)

# ============================================
# REPORT VIEWS - DARI REPORT MODEL
# ============================================

@login_required
def report(request):
    """Report page with filtering"""
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    search = request.GET.get('search')
    
    try:
        report_data = Report.objects.filter(report_type='Transaction').order_by('-created_at')
        
        if start_date:
            report_data = report_data.filter(created_at__date__gte=start_date)
        if end_date:
            report_data = report_data.filter(created_at__date__lte=end_date)
        if search:
            report_data = report_data.filter(
                Q(data__no_transaksi__icontains=search) |
                Q(data__barang__icontains=search) |
                Q(data__kustomer__icontains=search) |
                Q(data__supplier__icontains=search)
            )
        
        total_records = report_data.count()
        total_weight = 0
        weights = []
        
        for item in report_data:
            weight = item.data.get('weight', 0) if isinstance(item.data, dict) else 0
            if weight:
                try:
                    w = float(weight)
                    weights.append(w)
                except (ValueError, TypeError):
                    pass
        
        if weights:
            total_weight = sum(weights)
            avg_weight = sum(weights) / len(weights) if weights else 0
            max_weight = max(weights) if weights else 0
            min_weight = min(weights) if weights else 0
        else:
            total_weight = 0
            avg_weight = 0
            max_weight = 0
            min_weight = 0
        
        paginator = Paginator(report_data, 30)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)
    except:
        page_obj = []
        total_records = 0
        total_weight = 0
        avg_weight = 0
        max_weight = 0
        min_weight = 0
    
    return render(request, 'dashboard/report.html', {
        'page_obj': page_obj,
        'start_date': start_date,
        'end_date': end_date,
        'search': search,
        'total_weight': total_weight,
        'avg_weight': avg_weight,
        'max_weight': max_weight,
        'min_weight': min_weight,
        'total_records': total_records
    })

# ============================================
# REPORT DETAIL VIEW
# ============================================

# dashboard/views.py

@login_required
def report_detail(request, pk):
    """Get report detail for modal display"""
    try:
        report = Report.objects.get(id=pk, report_type='Transaction')
        data = report.data if isinstance(report.data, dict) else {}
        
        # ============================================
        # ✅ PERBAIKI URL FOTO
        # ============================================
        photo_url = ''
        file_path = getattr(report, 'file_path', '')
        file_path_str = str(file_path) if file_path else ''  # <-- DEFINISIKAN DI SINI
        
        if file_path and file_path_str != '' and file_path_str != 'None':
            # Cek apakah sudah berupa URL lengkap
            if file_path_str.startswith('http://') or file_path_str.startswith('https://'):
                photo_url = file_path_str
            # Cek apakah sudah ada /media/
            elif file_path_str.startswith('/media/'):
                photo_url = file_path_str
            else:
                # Tambahkan MEDIA_URL
                photo_url = settings.MEDIA_URL + file_path_str.lstrip('/')  # <-- AMAN
        
        # Debug
        print(f"📸 Report ID: {pk}")
        print(f"📸 file_path: {file_path}")
        print(f"📸 photo_url: {photo_url}")
        
        return JsonResponse({
            'status': 'success',
            'data': {
                'id': getattr(report, 'id', None),
                'no_transaksi': data.get('no_transaksi', ''),
                'barang': data.get('barang', ''),
                'barang_id': data.get('barang_id', ''),
                'barang_lot': data.get('barang_lot', ''),
                'kustomer': data.get('kustomer', ''),
                'kustomer_id': data.get('kustomer_id', ''),
                'supplier': data.get('supplier', ''),
                'supplier_id': data.get('supplier_id', ''),
                'weight': data.get('weight', 0),
                'keterangan': data.get('keterangan', ''),
                'created_at': report.created_at.strftime('%Y-%m-%d %H:%M:%S') if report.created_at else '',
                'operator': report.generated_by.username if report.generated_by else 'Unknown',
                'photo': photo_url
            }
        })
    except Report.DoesNotExist:
        return JsonResponse({
            'status': 'error',
            'message': 'Report not found'
        }, status=404)
    except Exception as e:
        print(f"❌ Report detail error: {e}")
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=500)
# ============================================
# REPORT DELETE VIEW
# ============================================

@login_required
def report_delete(request, pk):
    """Delete a report with FDA compliance"""
    if request.method != 'POST':
        return JsonResponse({
            'status': 'error',
            'message': 'Method not allowed'
        }, status=405)
    
    try:
        data = json.loads(request.body)
        password = data.get('password')
        
        if not password:
            return JsonResponse({
                'status': 'error', 
                'message': 'Password required for FDA compliance'
            }, status=400)
        
        if not request.user.check_password(password):
            return JsonResponse({
                'status': 'error', 
                'message': 'Wrong password'
            }, status=401)
        
        report = Report.objects.get(id=pk, report_type='Transaction')
        
        if report.file_path:
            try:
                path = os.path.join(settings.MEDIA_ROOT, str(report.file_path))
                if os.path.exists(path):
                    os.remove(path)
            except:
                pass
        
        try:
            UserActivity.objects.create(
                user=request.user,
                action='Delete Report',
                details=f'Deleted report ID: {pk}',
                ip_address=request.META.get('REMOTE_ADDR', '')
            )
        except:
            pass
        
        report.delete()
        
        return JsonResponse({
            'status': 'success',
            'message': 'Report deleted successfully'
        })
    except Report.DoesNotExist:
        return JsonResponse({
            'status': 'error',
            'message': 'Report not found'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=500)

# ============================================
# EXPORT REPORT EXCEL
# ============================================

@login_required
def export_report_excel(request):
    """Export report to Excel with FDA Part 11 compliance"""
    try:
        password = request.GET.get('password')
        if not password:
            return JsonResponse({
                'status': 'error', 
                'message': 'Password required for FDA Part 11 compliance'
            }, status=401)
        
        if not request.user.check_password(password):
            return JsonResponse({
                'status': 'error', 
                'message': 'Invalid password'
            }, status=401)
        
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        search = request.GET.get('search', '')
        ids_param = request.GET.get('ids', '')
        
        data = Report.objects.filter(report_type='Transaction').order_by('-created_at')
        
        if start_date:
            data = data.filter(created_at__date__gte=start_date)
        if end_date:
            data = data.filter(created_at__date__lte=end_date)
        if search:
            data = data.filter(
                Q(data__no_transaksi__icontains=search) |
                Q(data__barang__icontains=search) |
                Q(data__kustomer__icontains=search) |
                Q(data__supplier__icontains=search)
            )
        if ids_param:
            ids = [x for x in ids_param.split(',') if x]
            if ids:
                data = data.filter(id__in=ids)
        
        company = CompanyProfile.objects.first()
        company_name = company.name if company else 'PT Interskala Mandiri Indonesia'
        company_address = company.address if company else 'Green Sedayu Biz Park Jl. Daan Mogot KM. 18, Kalideres, Jakarta Barat'
        company_phone = company.phone if company else '(021) 2252-2992'
        company_email = company.email if company else 'sales@interskala.com'
        
        export_data = []
        error_messages = []
        
        for idx, item in enumerate(data):
            try:
                item_data = item.data if isinstance(item.data, dict) else {}
                export_data.append({
                    'No': idx + 1,
                    'No. Transaksi': item_data.get('no_transaksi', ''),
                    'Nama Barang': item_data.get('barang', ''),
                    'ID Barang': item_data.get('barang_id', ''),
                    'Lot': item_data.get('barang_lot', ''),
                    'Nama Kustomer': item_data.get('kustomer', ''),
                    'ID Kustomer': item_data.get('kustomer_id', ''),
                    'Nama Supplier': item_data.get('supplier', ''),
                    'ID Supplier': item_data.get('supplier_id', ''),
                    'Berat (kg)': item_data.get('weight', 0),
                    'Keterangan': item_data.get('keterangan', ''),
                    'Tanggal': item.created_at.strftime('%Y-%m-%d %H:%M:%S') if item.created_at else '',
                    'Operator': item.generated_by.username if item.generated_by else 'Unknown'
                })
            except Exception as e:
                error_messages.append(f"Error processing item {idx}: {str(e)}")
        
        df = pd.DataFrame(export_data)
        
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            metadata = pd.DataFrame({
                'Informasi Perusahaan': ['Nama Perusahaan', 'Alamat', 'Telepon', 'Email', 'FDA Compliance', 'Dicetak Oleh', 'Tanggal Cetak', 'Total Records'],
                'Nilai': [
                    company_name,
                    company_address,
                    company_phone,
                    company_email,
                    'FDA 21 CFR Part 11 Compliant',
                    request.user.get_full_name() or request.user.username,
                    timezone.now().strftime('%Y-%m-%d %H:%M:%S'),
                    len(data)
                ]
            })
            metadata.to_excel(writer, sheet_name='Metadata', index=False)
            df.to_excel(writer, sheet_name='Report', index=False)
            
            workbook = writer.book
            worksheet = writer.sheets['Report']
            
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            header_font = Font(bold=True, size=11, color='FFFFFF')
            header_fill = PatternFill(start_color='0D47A1', end_color='0D47A1', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center')
            
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
        
        try:
            UserActivity.objects.create(
                user=request.user,
                action='Export Excel',
                details=f'Exported {len(data)} records to Excel',
                ip_address=request.META.get('REMOTE_ADDR', '')
            )
        except:
            pass
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'report_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
        
    except Exception as e:
        logger.error(f"Export error: {e}", exc_info=True)
        return JsonResponse({
            'status': 'error',
            'message': f'Export failed: {str(e)}'
        }, status=500)

# ============================================
# DELETE REPORT ITEM
# ============================================

@login_required
def delete_report_item(request, item_id):
    """Delete single report item with FDA compliance"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            password = data.get('password')
            
            if not password:
                return JsonResponse({
                    'status': 'error', 
                    'message': 'Password required for FDA compliance'
                }, status=400)
            
            if not request.user.check_password(password):
                return JsonResponse({
                    'status': 'error', 
                    'message': 'Wrong password'
                }, status=401)
            
            item = get_object_or_404(Report, id=item_id, report_type='Transaction')
            
            if item.file_path:
                try:
                    path = os.path.join(settings.MEDIA_ROOT, str(item.file_path))
                    if os.path.exists(path):
                        os.remove(path)
                except:
                    pass
            
            try:
                UserActivity.objects.create(
                    user=request.user,
                    action='Delete Report Item',
                    details=f'Deleted report item ID: {item_id}',
                    ip_address=request.META.get('REMOTE_ADDR')
                )
            except:
                pass
            
            item.delete()
            
            return JsonResponse({
                'status': 'success', 
                'message': f'Item {item_id} deleted successfully'
            })
            
        except Exception as e:
            return JsonResponse({
                'status': 'error', 
                'message': str(e)
            }, status=400)
    
    return JsonResponse({
        'status': 'error', 
        'message': 'Invalid request method'
    }, status=405)

# ============================================
# EXPORT SELECTED EXCEL
# ============================================

@login_required
def export_selected_excel(request):
    """Export selected reports to Excel with FDA Part 11 compliance"""
    try:
        password = request.GET.get('password')
        if not password:
            return JsonResponse({
                'status': 'error', 
                'message': 'Password required for FDA Part 11 compliance'
            }, status=401)
        
        if not request.user.check_password(password):
            return JsonResponse({
                'status': 'error', 
                'message': 'Invalid password'
            }, status=401)
        
        ids_param = request.GET.get('ids', '')
        if not ids_param:
            return JsonResponse({
                'status': 'error',
                'message': 'No records selected'
            }, status=400)
        
        ids = [int(x) for x in ids_param.split(',') if x.isdigit()]
        if not ids:
            return JsonResponse({
                'status': 'error',
                'message': 'Invalid IDs provided'
            }, status=400)
        
        data = Report.objects.filter(
            id__in=ids,
            report_type='Transaction'
        ).order_by('-created_at')
        
        if not data.exists():
            return JsonResponse({
                'status': 'error',
                'message': 'No records found'
            }, status=404)
        
        company = CompanyProfile.objects.first()
        company_name = company.name if company else 'PT Interskala Mandiri Indonesia'
        company_address = company.address if company else 'Green Sedayu Biz Park Jl. Daan Mogot KM. 18, Kalideres, Jakarta Barat'
        company_phone = company.phone if company else '(021) 2252-2992'
        company_email = company.email if company else 'sales@interskala.com'
        
        export_data = []
        error_messages = []
        
        for idx, item in enumerate(data):
            try:
                item_data = item.data if isinstance(item.data, dict) else {}
                export_data.append({
                    'No': idx + 1,
                    'No. Transaksi': item_data.get('no_transaksi', ''),
                    'Nama Barang': item_data.get('barang', ''),
                    'ID Barang': item_data.get('barang_id', ''),
                    'Lot': item_data.get('barang_lot', ''),
                    'Nama Kustomer': item_data.get('kustomer', ''),
                    'ID Kustomer': item_data.get('kustomer_id', ''),
                    'Nama Supplier': item_data.get('supplier', ''),
                    'ID Supplier': item_data.get('supplier_id', ''),
                    'Berat (kg)': item_data.get('weight', 0),
                    'Keterangan': item_data.get('keterangan', ''),
                    'Tanggal': item.created_at.strftime('%Y-%m-%d %H:%M:%S') if item.created_at else '',
                    'Operator': item.generated_by.username if item.generated_by else 'Unknown'
                })
            except Exception as e:
                error_messages.append(f"Error processing item {idx}: {str(e)}")
        
        df = pd.DataFrame(export_data)
        
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            metadata = pd.DataFrame({
                'Informasi Perusahaan': ['Nama Perusahaan', 'Alamat', 'Telepon', 'Email', 'FDA Compliance', 'Dicetak Oleh', 'Tanggal Cetak', 'Total Records'],
                'Nilai': [
                    company_name,
                    company_address,
                    company_phone,
                    company_email,
                    'FDA 21 CFR Part 11 Compliant',
                    request.user.get_full_name() or request.user.username,
                    timezone.now().strftime('%Y-%m-%d %H:%M:%S'),
                    len(data)
                ]
            })
            metadata.to_excel(writer, sheet_name='Metadata', index=False)
            df.to_excel(writer, sheet_name='Report', index=False)
            
            workbook = writer.book
            worksheet = writer.sheets['Report']
            
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            header_font = Font(bold=True, size=11, color='FFFFFF')
            header_fill = PatternFill(start_color='0D47A1', end_color='0D47A1', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center')
            
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
        
        try:
            UserActivity.objects.create(
                user=request.user,
                action='Export Selected Excel',
                details=f'Exported {len(data)} selected records to Excel',
                ip_address=request.META.get('REMOTE_ADDR', '')
            )
        except:
            pass
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'selected_report_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
        
    except Exception as e:
        logger.error(f"Export selected error: {e}", exc_info=True)
        return JsonResponse({
            'status': 'error',
            'message': f'Export failed: {str(e)}'
        }, status=500)

# ============================================
# PRINT REPORT
# ============================================

@login_required
@csrf_exempt
def print_report(request):
    """Print report with FDA Part 11 compliance"""
    if request.method != 'POST':
        return JsonResponse({
            'status': 'error',
            'message': 'Method not allowed'
        }, status=405)
    
    try:
        data = json.loads(request.body)
        password = data.get('password')
        ids = data.get('ids', [])
        save_after_print = data.get('save', True)
        
        if not password:
            return JsonResponse({
                'status': 'error',
                'message': 'Password required for FDA Part 11 compliance'
            }, status=401)
        
        if not request.user.check_password(password):
            return JsonResponse({
                'status': 'error',
                'message': 'Invalid password'
            }, status=401)
        
        if ids:
            reports = Report.objects.filter(
                id__in=ids,
                report_type='Transaction'
            ).order_by('-created_at')
        else:
            reports = Report.objects.filter(
                report_type='Transaction'
            ).order_by('-created_at')
        
        if not reports.exists():
            return JsonResponse({
                'status': 'error',
                'message': 'No records found'
            }, status=404)
        
        company = CompanyProfile.objects.first()
        
        print_data = []
        for idx, item in enumerate(reports):
            item_data = item.data if isinstance(item.data, dict) else {}
            photo_url = get_photo_url(item.file_path)
            
            print_data.append({
                'no': idx + 1,
                'no_transaksi': item_data.get('no_transaksi', ''),
                'barang': item_data.get('barang', ''),
                'barang_id': item_data.get('barang_id', ''),
                'lot': item_data.get('barang_lot', ''),
                'kustomer': item_data.get('kustomer', ''),
                'kustomer_id': item_data.get('kustomer_id', ''),
                'supplier': item_data.get('supplier', ''),
                'supplier_id': item_data.get('supplier_id', ''),
                'weight': item_data.get('weight', 0),
                'keterangan': item_data.get('keterangan', ''),
                'tanggal': item.created_at.strftime('%Y-%m-%d %H:%M:%S') if item.created_at else '',
                'operator': item.generated_by.username if item.generated_by else 'Unknown',
                'photo': photo_url,
                'report_id': getattr(item, 'id', None)
            })
        
        saved_count = 0
        if save_after_print:
            saved_count = len(reports)
            for report in reports:
                try:
                    UserActivity.objects.create(
                        user=request.user,
                        action='Print & Save Report',
                        details=f'Printed and saved report ID: {getattr(report, "id", "Unknown")}',
                        ip_address=request.META.get('REMOTE_ADDR', '')
                    )
                except:
                    pass
        
        try:
            UserActivity.objects.create(
                user=request.user,
                action='Print Report',
                details=f'Printed {len(print_data)} records. Saved: {save_after_print}',
                ip_address=request.META.get('REMOTE_ADDR', '')
            )
        except:
            pass
        
        return JsonResponse({
            'status': 'success',
            'data': print_data,
            'company': {
                'name': company.name if company else 'PT Interskala Mandiri Indonesia',
                'address': company.address if company else 'Green Sedayu Biz Park Jl. Daan Mogot KM. 18, Kalideres, Jakarta Barat',
                'phone': company.phone if company else '(021) 2252-2992',
                'email': company.email if company else 'sales@interskala.com'
            },
            'printed_by': request.user.get_full_name() or request.user.username,
            'printed_at': timezone.now().strftime('%Y-%m-%d %H:%M:%S'),
            'total_records': len(print_data),
            'saved': save_after_print,
            'saved_count': saved_count
        })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'status': 'error',
            'message': 'Invalid JSON data'
        }, status=400)
    except Exception as e:
        logger.error(f"Print error: {e}", exc_info=True)
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=500)

# ============================================
# CAPTURE / SAVE VIEWS
# ============================================

@login_required
def save_capture(request):
    """Save captured weight data"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            weight = data.get('weight')
            latitude = data.get('latitude')
            longitude = data.get('longitude')
            
            if weight is None:
                return JsonResponse({'status': 'error', 'message': 'Weight data required'})
            
            new_entry = WeightData.objects.create(
                entry_id=WeightData.objects.count() + 1,
                weight=float(weight),
                latitude=float(latitude) if latitude else -6.152256,
                longitude=float(longitude) if longitude else 106.694091,
                created_at=timezone.now()
            )
            
            try:
                UserActivity.objects.create(
                    user=request.user,
                    action='Capture',
                    details=f'Captured weight: {weight} kg',
                    ip_address=request.META.get('REMOTE_ADDR')
                )
            except:
                pass
            
            return JsonResponse({
                'status': 'success',
                'message': 'Data saved successfully',
                'data': {
                    'entry_id': new_entry.entry_id,
                    'weight': str(new_entry.weight),
                    'created_at': new_entry.created_at.strftime('%Y-%m-%d %H:%M:%S')
                }
            })
            
        except json.JSONDecodeError:
            return JsonResponse({'status': 'error', 'message': 'Invalid JSON data'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': f'Error: {str(e)}'})
    
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'})

@login_required
@csrf_exempt
def upload_report_photo(request):
    """Upload foto untuk report"""
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Method not allowed'}, status=405)
    
    try:
        # Cek permission
        if not request.user.is_staff and not request.user.is_superuser:
            return JsonResponse({'status': 'error', 'message': 'Permission denied'}, status=403)
        
        # Ambil data dari request
        report_id = request.POST.get('report_id')
        photo = request.FILES.get('photo')
        
        if not report_id:
            return JsonResponse({'status': 'error', 'message': 'Report ID required'}, status=400)
        
        if not photo:
            return JsonResponse({'status': 'error', 'message': 'No photo uploaded'}, status=400)
        
        # Validasi file
        allowed_types = ['image/jpeg', 'image/png', 'image/gif', 'image/webp']
        if photo.content_type not in allowed_types:
            return JsonResponse({'status': 'error', 'message': 'Invalid file type. Use JPG, PNG, GIF, or WEBP'}, status=400)
        
        # Dapatkan report
        report = get_object_or_404(Report, id=report_id, report_type='Transaction')
        
        # Buat nama file unik
        ext = os.path.splitext(photo.name)[1]
        filename = f"report_{report_id}_{uuid.uuid4().hex[:8]}{ext}"
        filepath = os.path.join('reports', filename)
        
        # Simpan file
        full_path = os.path.join(settings.MEDIA_ROOT, filepath)
        os.makedirs(os.path.dirname(full_path), exist_ok=True)
        
        with open(full_path, 'wb+') as destination:
            for chunk in photo.chunks():
                destination.write(chunk)
        
        # Hapus foto lama jika ada
        if report.file_path and os.path.exists(os.path.join(settings.MEDIA_ROOT, report.file_path)):
            try:
                os.remove(os.path.join(settings.MEDIA_ROOT, report.file_path))
            except:
                pass
        
        # Update report dengan file_path
        report.file_path = filepath
        report.save()
        
        # Log aktivitas
        try:
            UserActivity.objects.create(
                user=request.user,
                action='Upload Photo',
                details=f'Uploaded photo for report ID: {report_id}',
                ip_address=request.META.get('REMOTE_ADDR', '')
            )
        except:
            pass
        
        return JsonResponse({
            'status': 'success',
            'message': 'Photo uploaded successfully',
            'file_path': filepath,
            'url': settings.MEDIA_URL + filepath,
            'report_id': report_id
        })
        
    except Report.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Report not found'}, status=404)
    except Exception as e:
        print(f"❌ Upload photo error: {e}")
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
    
# ============================================
# TRANSACTION VIEWS - SAVE KE REPORT
# ============================================
@login_required
def save_transaction(request):
    """Save transaction to Report dengan foto disimpan sebagai file"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            
            barang_name = data.get('barang', '')
            barang_id = data.get('barang_id', '')
            barang_lot = data.get('barang_lot', '')
            kustomer_name = data.get('kustomer', '')
            kustomer_id = data.get('kustomer_id', '')
            supplier_name = data.get('supplier', '')
            supplier_id = data.get('supplier_id', '')
            no_transaksi = data.get('no_transaksi', '')
            weight = data.get('weight', 0)
            keterangan = data.get('keterangan', '')
            photo_data = data.get('photo', '')  # Base64 image
            
            # ============================================
            # ✅ SIMPAN FOTO SEBAGAI FILE, BUKAN DI DATABASE
            # ============================================
            file_path = ''
            if photo_data and photo_data.startswith('data:image'):
                try:
                    # Ekstrak base64
                    import base64
                    format, imgstr = photo_data.split(';base64,')
                    ext = format.split('/')[-1]
                    
                    # Buat nama file unik
                    filename = f"report_{no_transaksi}_{uuid.uuid4().hex[:8]}.{ext}"
                    filepath = os.path.join('reports', filename)
                    
                    # Simpan ke media folder
                    full_path = os.path.join(settings.MEDIA_ROOT, filepath)
                    os.makedirs(os.path.dirname(full_path), exist_ok=True)
                    
                    with open(full_path, 'wb') as f:
                        f.write(base64.b64decode(imgstr))
                    
                    file_path = filepath  # Simpan path saja, bukan base64
                    print(f"✅ Photo saved: {file_path}")
                    
                except Exception as e:
                    print(f"❌ Save photo error: {e}")
                    file_path = ''
            
            # ============================================
            # ✅ CEK REPORT EXISTING
            # ============================================
            existing_report = Report.objects.filter(
                report_type='Transaction',
                data__no_transaksi=no_transaksi
            ).first()
            
            if existing_report:
                # UPDATE
                report_data = existing_report.data if isinstance(existing_report.data, dict) else {}
                report_data.update({
                    'no_transaksi': no_transaksi,
                    'barang': barang_name,
                    'barang_id': barang_id,
                    'barang_lot': barang_lot,
                    'kustomer': kustomer_name,
                    'kustomer_id': kustomer_id,
                    'supplier': supplier_name,
                    'supplier_id': supplier_id,
                    'weight': str(weight),
                    'keterangan': keterangan,
                })
                existing_report.data = report_data
                
                if file_path:
                    existing_report.file_path = file_path
                
                existing_report.save()
                
                return JsonResponse({
                    'status': 'success',
                    'message': 'Transaction updated successfully',
                    'id': existing_report.pk,
                    'no_transaksi': no_transaksi,
                    'weight': str(weight),
                    'updated': True
                })
            
            # ============================================
            # ✅ CREATE NEW REPORT
            # ============================================
            report_data = {
                'no_transaksi': no_transaksi,
                'barang': barang_name,
                'barang_id': barang_id,
                'barang_lot': barang_lot,
                'kustomer': kustomer_name,
                'kustomer_id': kustomer_id,
                'supplier': supplier_name,
                'supplier_id': supplier_id,
                'weight': str(weight),
                'keterangan': keterangan,
                'created_at': timezone.now().isoformat()
            }
            
            transaction = Report.objects.create(
                report_type='Transaction',
                generated_by=request.user,
                start_date=timezone.now(),
                end_date=timezone.now(),
                data=report_data,
                file_path=file_path  # ✅ HANYA PATH, BUKAN BASE64
            )
            
            try:
                UserActivity.objects.create(
                    user=request.user,
                    action='Save Transaction',
                    details=f'Transaction saved: {no_transaksi}',
                    ip_address=request.META.get('REMOTE_ADDR')
                )
            except:
                pass
            
            return JsonResponse({
                'status': 'success',
                'message': 'Transaction saved successfully',
                'id': transaction.pk,
                'no_transaksi': no_transaksi,
                'weight': str(weight),
                'updated': False
            })
            
        except json.JSONDecodeError as e:
            return JsonResponse({'status': 'error', 'message': f'Invalid JSON: {str(e)}'})
        except Exception as e:
            logger.error(f"Save transaction error: {e}", exc_info=True)
            return JsonResponse({'status': 'error', 'message': str(e)})
    
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'})

# ============================================
# DELETE SELECTED RECORDS
# ============================================

@login_required
def delete_selected_records(request):
    """Delete selected records with FDA compliance"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            ids = data.get('ids', [])
            password = data.get('password')
            
            if not password:
                return JsonResponse({'status': 'error', 'message': 'Password required for FDA compliance'})
            
            if not request.user.check_password(password):
                return JsonResponse({'status': 'error', 'message': 'Wrong password'})
            
            if not ids:
                return JsonResponse({'status': 'error', 'message': 'No records selected'})
            
            records = Report.objects.filter(id__in=ids, report_type='Transaction')
            count = records.count()
            
            for record in records:
                if record.file_path:
                    try:
                        path = os.path.join(settings.MEDIA_ROOT, str(record.file_path))
                        if os.path.exists(path):
                            os.remove(path)
                    except:
                        pass
            
            try:
                UserActivity.objects.create(
                    user=request.user,
                    action='Delete Selected Records',
                    details=f'Deleted {count} records from report',
                    ip_address=request.META.get('REMOTE_ADDR')
                )
            except:
                pass
            
            records.delete()
            
            return JsonResponse({
                'status': 'success',
                'message': f'{count} records deleted successfully'
            })
            
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})
    
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'})

# ============================================
# DELETE ALL RECORDS
# ============================================

@login_required
def delete_all_records(request):
    """Delete all records with FDA compliance"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            password = data.get('password')
            
            if not password:
                return JsonResponse({'status': 'error', 'message': 'Password required for FDA compliance'})
            
            if not request.user.check_password(password):
                return JsonResponse({'status': 'error', 'message': 'Wrong password'})
            
            records = Report.objects.filter(report_type='Transaction')
            count = records.count()
            
            for record in records:
                if record.file_path:
                    try:
                        path = os.path.join(settings.MEDIA_ROOT, str(record.file_path))
                        if os.path.exists(path):
                            os.remove(path)
                    except:
                        pass
            
            try:
                UserActivity.objects.create(
                    user=request.user,
                    action='Delete All Records',
                    details=f'Deleted ALL {count} records from report',
                    ip_address=request.META.get('REMOTE_ADDR')
                )
            except:
                pass
            
            records.delete()
            
            return JsonResponse({
                'status': 'success',
                'message': f'All {count} records deleted successfully'
            })
            
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})
    
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'})

# ============================================
# CRUD OPERATIONS - BARANG
# ============================================

@login_required
def get_barang(request, id):
    """Get barang data with password verification for editing"""
    try:
        password = request.GET.get('password')
        
        if not password:
            return JsonResponse({
                'status': 'error',
                'message': 'Password required for FDA compliance'
            }, status=401)
        
        if not request.user.check_password(password):
            return JsonResponse({
                'status': 'error',
                'message': 'Invalid password'
            }, status=401)
        
        id_barang = str(id).strip()
        
        try:
            barang = Barang.objects.get(id_barang__iexact=id_barang)
        except Barang.DoesNotExist:
            barang = Barang.objects.filter(id_barang__icontains=id_barang).first()
            if not barang:
                return JsonResponse({
                    'status': 'error',
                    'message': f'Barang with ID "{id}" not found'
                }, status=404)
        
        data = {
            'id_barang': barang.id_barang,
            'nama_barang': barang.nama_barang,
            'lot': barang.lot,
            'kategori': barang.kategori,
            'deskripsi': barang.deskripsi,
        }
        
        try:
            UserActivity.objects.create(
                user=request.user,
                action='View Barang',
                details=f'Viewed barang: {id_barang} for editing',
                ip_address=request.META.get('REMOTE_ADDR')
            )
        except:
            pass
        
        return JsonResponse({'status': 'success', 'data': data})
        
    except Exception as e:
        logger.error(f"Get barang error: {e}")
        return JsonResponse({
            'status': 'error', 
            'message': str(e)
        }, status=404)

@login_required
@csrf_exempt
def update_barang(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            
            # LOG UNTUK DEBUG
            print("=" * 50)
            print("UPDATE BARANG REQUEST")
            print(f"Data keys: {list(data.keys())}")
            print(f"Data: {data}")
            print(f"User: {request.user.username}")
            print("=" * 50)
            
            id_barang = data.get('id_barang')
            password = data.get('password')
            
            # COBA AMBIL PASSWORD DARI BEBERAPA KEY
            password = data.get('password') or data.get('edit_password') or data.get('pwd') or data.get('pass')
            
            if not password:
                return JsonResponse({
                    'status': 'error', 
                    'message': 'Password required for FDA compliance. Please enter your password.'
                }, status=401)
            
            if not request.user.check_password(password):
                logger.warning(f"Wrong password for user: {request.user.username}")
                return JsonResponse({
                    'status': 'error', 
                    'message': 'Wrong password. Please try again.'
                }, status=401)
            
            if not id_barang:
                return JsonResponse({
                    'status': 'error',
                    'message': 'ID Barang is required'
                }, status=400)
            
            id_barang = str(id_barang).strip()
            
            try:
                barang = Barang.objects.get(id_barang__iexact=id_barang)
            except Barang.DoesNotExist:
                barang = Barang.objects.filter(id_barang__icontains=id_barang).first()
                if not barang:
                    return JsonResponse({
                        'status': 'error',
                        'message': f'Barang with ID "{id_barang}" not found'
                    }, status=404)
            
            barang.nama_barang = data.get('nama_barang')
            barang.lot = data.get('lot')
            barang.kategori = data.get('kategori', '')
            barang.deskripsi = data.get('deskripsi', '')
            barang.save()
            
            try:
                UserActivity.objects.create(
                    user=request.user,
                    action='Update Barang',
                    details=f'Updated barang: {id_barang}',
                    ip_address=request.META.get('REMOTE_ADDR')
                )
            except:
                pass
            
            return JsonResponse({
                'status': 'success',
                'message': f'Barang {id_barang} updated successfully'
            })
            
        except json.JSONDecodeError as e:
            logger.error(f"JSON decode error: {e}")
            return JsonResponse({'status': 'error', 'message': 'Invalid JSON data'}, status=400)
        except Exception as e:
            logger.error(f"Update barang error: {e}", exc_info=True)
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    
    return JsonResponse({'status': 'error', 'message': 'Method not allowed'}, status=405)

@login_required
def delete_barang(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            
            print("=" * 50)
            print("DELETE BARANG REQUEST")
            print(f"User: {request.user.username}")
            print(f"Data keys: {list(data.keys())}")
            print("=" * 50)
            
            id_barang = data.get('id')
            password = data.get('password')
            
            # Coba cari password di key lain
            if not password:
                password = data.get('edit_password') or data.get('pwd') or data.get('pass')
            
            if not password:
                return JsonResponse({
                    'status': 'error', 
                    'message': 'Password required for FDA compliance. Please enter your password.'
                }, status=401)
            
            if not request.user.check_password(password):
                return JsonResponse({
                    'status': 'error', 
                    'message': 'Wrong password. Please try again.'
                }, status=401)
            
            if not id_barang:
                return JsonResponse({
                    'status': 'error',
                    'message': 'ID Barang is required'
                }, status=400)
            
            barang = get_object_or_404(Barang, id_barang=id_barang)
            nama_barang = barang.nama_barang
            
            try:
                UserActivity.objects.create(
                    user=request.user,
                    action='Delete Barang',
                    details=f'Deleted barang: {id_barang} - {nama_barang}',
                    ip_address=request.META.get('REMOTE_ADDR')
                )
            except:
                pass
            
            barang.delete()
            
            return JsonResponse({
                'status': 'success',
                'message': f'Barang {id_barang} deleted successfully'
            })
            
        except Barang.DoesNotExist:
            return JsonResponse({
                'status': 'error',
                'message': f'Barang with ID {id_barang} not found'
            }, status=404)
        except json.JSONDecodeError as e:
            return JsonResponse({'status': 'error', 'message': 'Invalid JSON data'}, status=400)
        except Exception as e:
            logger.error(f"Delete barang error: {e}", exc_info=True)
            return JsonResponse({
                'status': 'error', 
                'message': str(e)
            }, status=400)
    
    return JsonResponse({
        'status': 'error', 
        'message': 'Method not allowed'
    }, status=405)
    
# ============================================
# CRUD OPERATIONS - KUSTOMER
# ============================================

@login_required
def get_kustomer(request, id):
    """Get kustomer data with password verification for editing"""
    try:
        password = request.GET.get('password')
        
        if not password:
            return JsonResponse({
                'status': 'error',
                'message': 'Password required for FDA compliance'
            }, status=401)
        
        if not request.user.check_password(password):
            return JsonResponse({
                'status': 'error',
                'message': 'Invalid password'
            }, status=401)
        
        id_kustomer = str(id).strip()
        
        try:
            kustomer = Kustomer.objects.get(id_kustomer__iexact=id_kustomer)
        except Kustomer.DoesNotExist:
            kustomer = Kustomer.objects.filter(id_kustomer__icontains=id_kustomer).first()
            if not kustomer:
                return JsonResponse({
                    'status': 'error',
                    'message': f'Kustomer with ID "{id}" not found'
                }, status=404)
        
        data = {
            'id_kustomer': kustomer.id_kustomer,
            'nama_kustomer': kustomer.nama_kustomer,
            'alamat': kustomer.alamat,
            'telepon': kustomer.telepon,
            'email': kustomer.email,
        }
        
        try:
            UserActivity.objects.create(
                user=request.user,
                action='View Kustomer',
                details=f'Viewed kustomer: {id_kustomer} for editing',
                ip_address=request.META.get('REMOTE_ADDR')
            )
        except:
            pass
        
        return JsonResponse({'status': 'success', 'data': data})
        
    except Exception as e:
        logger.error(f"Get kustomer error: {e}")
        return JsonResponse({
            'status': 'error', 
            'message': str(e)
        }, status=404)

@login_required
def update_kustomer(request):
    """Update kustomer with FDA compliance"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            id_kustomer = data.get('id_kustomer')
            password = data.get('password')
            
            if not password:
                return JsonResponse({
                    'status': 'error', 
                    'message': 'Password required for FDA compliance'
                }, status=401)
            
            if not request.user.check_password(password):
                return JsonResponse({
                    'status': 'error', 
                    'message': 'Wrong password'
                }, status=401)
            
            if not id_kustomer:
                return JsonResponse({
                    'status': 'error',
                    'message': 'ID Kustomer is required'
                }, status=400)
            
            id_kustomer = str(id_kustomer).strip()
            
            try:
                kustomer = Kustomer.objects.get(id_kustomer__iexact=id_kustomer)
            except Kustomer.DoesNotExist:
                kustomer = Kustomer.objects.filter(id_kustomer__icontains=id_kustomer).first()
                if not kustomer:
                    return JsonResponse({
                        'status': 'error',
                        'message': f'Kustomer with ID "{id_kustomer}" not found'
                    }, status=404)
            
            kustomer.nama_kustomer = data.get('nama_kustomer')
            kustomer.alamat = data.get('alamat')
            kustomer.telepon = data.get('telepon')
            kustomer.email = data.get('email', '')
            kustomer.save()
            
            try:
                UserActivity.objects.create(
                    user=request.user,
                    action='Update Kustomer',
                    details=f'Updated kustomer: {id_kustomer}',
                    ip_address=request.META.get('REMOTE_ADDR')
                )
            except:
                pass
            
            return JsonResponse({
                'status': 'success',
                'message': f'Kustomer {id_kustomer} updated successfully'
            })
            
        except Exception as e:
            logger.error(f"Update kustomer error: {e}")
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    
    return JsonResponse({'status': 'error', 'message': 'Method not allowed'}, status=405)

@login_required
def delete_kustomer(request):
    """Delete kustomer with FDA compliance"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            id_kustomer = data.get('id')
            password = data.get('password')
            
            if not password:
                return JsonResponse({
                    'status': 'error', 
                    'message': 'Password required for FDA compliance'
                }, status=400)
            
            if not request.user.check_password(password):
                return JsonResponse({
                    'status': 'error', 
                    'message': 'Wrong password'
                }, status=401)
            
            if not id_kustomer:
                return JsonResponse({
                    'status': 'error',
                    'message': 'ID Kustomer is required'
                }, status=400)
            
            kustomer = get_object_or_404(Kustomer, id_kustomer=id_kustomer)
            nama_kustomer = kustomer.nama_kustomer
            
            try:
                UserActivity.objects.create(
                    user=request.user,
                    action='Delete Kustomer',
                    details=f'Deleted kustomer: {id_kustomer} - {nama_kustomer}',
                    ip_address=request.META.get('REMOTE_ADDR')
                )
            except:
                pass
            
            kustomer.delete()
            
            return JsonResponse({
                'status': 'success',
                'message': f'Kustomer {id_kustomer} deleted successfully'
            })
            
        except Kustomer.DoesNotExist:
            return JsonResponse({
                'status': 'error',
                'message': f'Kustomer with ID {id_kustomer} not found'
            }, status=404)
        except Exception as e:
            return JsonResponse({
                'status': 'error', 
                'message': str(e)
            }, status=400)
    
    return JsonResponse({
        'status': 'error', 
        'message': 'Method not allowed'
    }, status=405)

# ============================================
# CRUD OPERATIONS - SUPPLIER
# ============================================

@login_required
def get_supplier(request, id):
    """Get supplier data with password verification for editing"""
    try:
        password = request.GET.get('password')
        
        if not password:
            return JsonResponse({
                'status': 'error',
                'message': 'Password required for FDA compliance'
            }, status=401)
        
        if not request.user.check_password(password):
            return JsonResponse({
                'status': 'error',
                'message': 'Invalid password'
            }, status=401)
        
        id_supplier = str(id).strip()
        
        try:
            supplier = Supplier.objects.get(id_supplier__iexact=id_supplier)
        except Supplier.DoesNotExist:
            supplier = Supplier.objects.filter(id_supplier__icontains=id_supplier).first()
            if not supplier:
                return JsonResponse({
                    'status': 'error',
                    'message': f'Supplier with ID "{id}" not found'
                }, status=404)
        
        data = {
            'id_supplier': supplier.id_supplier,
            'nama_supplier': supplier.nama_supplier,
            'alamat': supplier.alamat,
            'telepon': supplier.telepon,
            'email': supplier.email,
        }
        
        try:
            UserActivity.objects.create(
                user=request.user,
                action='View Supplier',
                details=f'Viewed supplier: {id_supplier} for editing',
                ip_address=request.META.get('REMOTE_ADDR')
            )
        except:
            pass
        
        return JsonResponse({'status': 'success', 'data': data})
        
    except Exception as e:
        logger.error(f"Get supplier error: {e}")
        return JsonResponse({
            'status': 'error', 
            'message': str(e)
        }, status=404)

@login_required
def update_supplier(request):
    """Update supplier with FDA compliance"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            id_supplier = data.get('id_supplier')
            password = data.get('password')
            
            if not password:
                return JsonResponse({
                    'status': 'error', 
                    'message': 'Password required for FDA compliance'
                }, status=401)
            
            if not request.user.check_password(password):
                return JsonResponse({
                    'status': 'error', 
                    'message': 'Wrong password'
                }, status=401)
            
            if not id_supplier:
                return JsonResponse({
                    'status': 'error',
                    'message': 'ID Supplier is required'
                }, status=400)
            
            id_supplier = str(id_supplier).strip()
            
            try:
                supplier = Supplier.objects.get(id_supplier__iexact=id_supplier)
            except Supplier.DoesNotExist:
                supplier = Supplier.objects.filter(id_supplier__icontains=id_supplier).first()
                if not supplier:
                    return JsonResponse({
                        'status': 'error',
                        'message': f'Supplier with ID "{id_supplier}" not found'
                    }, status=404)
            
            supplier.nama_supplier = data.get('nama_supplier')
            supplier.alamat = data.get('alamat')
            supplier.telepon = data.get('telepon')
            supplier.email = data.get('email', '')
            supplier.save()
            
            try:
                UserActivity.objects.create(
                    user=request.user,
                    action='Update Supplier',
                    details=f'Updated supplier: {id_supplier}',
                    ip_address=request.META.get('REMOTE_ADDR')
                )
            except:
                pass
            
            return JsonResponse({
                'status': 'success',
                'message': f'Supplier {id_supplier} updated successfully'
            })
            
        except Exception as e:
            logger.error(f"Update supplier error: {e}")
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    
    return JsonResponse({'status': 'error', 'message': 'Method not allowed'}, status=405)

@login_required
def delete_supplier(request):
    """Delete supplier with FDA compliance"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            id_supplier = data.get('id')
            password = data.get('password')
            
            if not password:
                return JsonResponse({
                    'status': 'error', 
                    'message': 'Password required for FDA compliance'
                }, status=400)
            
            if not request.user.check_password(password):
                return JsonResponse({
                    'status': 'error', 
                    'message': 'Wrong password'
                }, status=401)
            
            if not id_supplier:
                return JsonResponse({
                    'status': 'error',
                    'message': 'ID Supplier is required'
                }, status=400)
            
            supplier = get_object_or_404(Supplier, id_supplier=id_supplier)
            nama_supplier = supplier.nama_supplier
            
            try:
                UserActivity.objects.create(
                    user=request.user,
                    action='Delete Supplier',
                    details=f'Deleted supplier: {id_supplier} - {nama_supplier}',
                    ip_address=request.META.get('REMOTE_ADDR')
                )
            except:
                pass
            
            supplier.delete()
            
            return JsonResponse({
                'status': 'success',
                'message': f'Supplier {id_supplier} deleted successfully'
            })
            
        except Supplier.DoesNotExist:
            return JsonResponse({
                'status': 'error',
                'message': f'Supplier with ID {id_supplier} not found'
            }, status=404)
        except Exception as e:
            return JsonResponse({
                'status': 'error', 
                'message': str(e)
            }, status=400)
    
    return JsonResponse({
        'status': 'error', 
        'message': 'Method not allowed'
    }, status=405)

# ============================================
# GET USER - DIPERBAIKI UNTUK CASE-INSENSITIVE
# ============================================

@login_required
def get_user(request, username):
    """Get user details - case insensitive"""
    try:
        print(f"🔍 get_user called with username: '{username}'")
        
        # Cari user dengan case-insensitive
        user = User.objects.get(username__iexact=username)
        
        data = {
            'username': user.username,
            'email': user.email,
            'is_superuser': user.is_superuser,
            'is_staff': user.is_staff,
            'is_active': user.is_active,
        }
        
        print(f"✅ Found user: {data}")
        return JsonResponse({'status': 'success', 'data': data})
        
    except User.DoesNotExist:
        print(f"❌ User not found: '{username}'")
        return JsonResponse({
            'status': 'error',
            'message': f'User "{username}" not found'
        }, status=404)
    except Exception as e:
        print(f"❌ Error: {str(e)}")
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=404)


# ============================================
# UPDATE USER - DIPERBAIKI DENGAN VERIFY PASSWORD (FDA Part 11)
# ============================================
@login_required
def update_user(request):
    """Update user with FDA compliance - verify_password required"""
    if request.method != 'POST':
        return JsonResponse({
            'status': 'error',
            'message': 'Method not allowed'
        }, status=405)
    
    try:
        data = json.loads(request.body)
        
        # Ambil data dari request
        username = data.get('username')
        email = data.get('email', '')
        password = data.get('password', '')
        is_superuser = data.get('is_superuser', False)
        is_staff = data.get('is_staff', False)
        verify_password = data.get('verify_password', '')
        
        # ✅ DEBUG - Log semua data
        print("=" * 60)
        print("📝 UPDATE USER REQUEST")
        print(f"  Username: {username}")
        print(f"  Email: {email}")
        print(f"  Password: {'[SET]' if password else '[EMPTY]'}")
        print(f"  Is Superuser: {is_superuser}")
        print(f"  Is Staff: {is_staff}")
        print(f"  Verify Password: {'[FILLED]' if verify_password else '[EMPTY]'}")
        print(f"  Logged in as: {request.user.username}")
        print("=" * 60)
        
        # ✅ VALIDASI: Username harus ada
        if not username:
            return JsonResponse({
                'status': 'error',
                'message': 'Username is required'
            }, status=400)
        
        # ✅ VALIDASI: Password verifikasi wajib diisi (FDA Part 11)
        if not verify_password:
            return JsonResponse({
                'status': 'error',
                'message': 'Password verifikasi diperlukan sesuai FDA 21 CFR Part 11'
            }, status=400)
        
        # ✅ CEK PASSWORD VERIFIKASI
        # ⚠️ PERHATIAN: Ini membandingkan dengan password user yang LOGIN (request.user)
        is_valid = request.user.check_password(verify_password)
        print(f"  Password verification result: {is_valid}")
        
        if not is_valid:
            return JsonResponse({
                'status': 'error',
                'message': 'Password verifikasi salah. Silakan coba lagi.'
            }, status=401)
        
        # ✅ Cari user yang akan diupdate
        try:
            user = User.objects.get(username__iexact=username)
            print(f"  Found user: {user.username}")
        except User.DoesNotExist:
            return JsonResponse({
                'status': 'error',
                'message': f'User "{username}" tidak ditemukan'
            }, status=404)
        
        # ✅ Update email
        if email:
            user.email = email
        
        # ✅ Update password jika diisi
        if password and password.strip():
            user.set_password(password)
            print(f"  Password updated for: {username}")
        
        # ✅ Update role (hanya superadmin)
        if request.user.is_superuser:
            user.is_superuser = bool(is_superuser)
            user.is_staff = bool(is_staff) or bool(is_superuser)
            print(f"  Role updated: Superuser={user.is_superuser}, Staff={user.is_staff}")
        else:
            print(f"  User {request.user.username} is not superadmin, skipping role update")
        
        user.save()
        
        # ✅ Log aktivitas
        try:
            UserActivity.objects.create(
                user=request.user,
                action='Update User',
                details=f'Updated user: {username}',
                ip_address=request.META.get('REMOTE_ADDR', '')
            )
        except Exception as e:
            print(f"  Logging error: {e}")
        
        return JsonResponse({
            'status': 'success',
            'message': f'User {username} updated successfully'
        })
        
    except json.JSONDecodeError as e:
        return JsonResponse({
            'status': 'error',
            'message': f'Invalid JSON: {str(e)}'
        }, status=400)
    except Exception as e:
        print(f"❌ Update user error: {e}")
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=500)
        
# ============================================
# API LATEST - UNTUK DASHBOARD REALTIME
# ============================================

@login_required
def api_latest(request):
    """API untuk mendapatkan data terbaru - DENGAN FILTER DATA BASI"""
    try:
        now = timezone.now()
        
        # ✅ Coba dari database dulu
        latest = WeightData.objects.all().order_by('-created_at').first()
        
        if latest:
            weight = float(latest.weight)
            diff = (now - latest.created_at).total_seconds() if latest.created_at else 999
            
            # ✅ JIKA DATA DI DATABASE LEBIH DARI 30 DETIK, AMBIL DARI THINGSPEAK
            if diff > 30:
                api_setting = APISetting.objects.first()
                if api_setting and api_setting.channel_id and api_setting.read_api_key:
                    try:
                        url = f"https://api.thingspeak.com/channels/{api_setting.channel_id}/feeds/last.json"
                        params = {'api_key': api_setting.read_api_key}
                        response = requests.get(url, params=params, timeout=5)
                        if response.status_code == 200:
                            feed = response.json()
                            if feed:
                                ts_weight = float(feed.get('field1', 0))
                                ts_created = feed.get('created_at')
                                
                                if ts_created:
                                    try:
                                        ts_time = datetime.strptime(ts_created, '%Y-%m-%dT%H:%M:%SZ')
                                        ts_time = timezone.make_aware(ts_time)
                                        ts_diff = (now - ts_time).total_seconds()
                                        
                                        # ✅ JIKA DATA THINGSPEAK LEBIH BARU (< 30 detik), PAKAI ITU
                                        if ts_diff < 30:
                                            weight = ts_weight
                                            # Simpan ke database
                                            if weight > 0:
                                                WeightData.objects.update_or_create(
                                                    entry_id=str(feed.get('entry_id')),
                                                    defaults={
                                                        'weight': weight,
                                                        'latitude': float(feed.get('field2', -6.152256)),
                                                        'longitude': float(feed.get('field3', 106.694091)),
                                                        'created_at': ts_time
                                                    }
                                                )
                                            else:
                                                # ✅ JIKA WEIGHT = 0, BUAT ATAU UPDATE DENGAN DATA BARU
                                                WeightData.objects.update_or_create(
                                                    entry_id=str(feed.get('entry_id')) if feed.get('entry_id') else '0',
                                                    defaults={
                                                        'weight': 0.00,
                                                        'latitude': -6.152256,
                                                        'longitude': 106.694091,
                                                        'created_at': ts_time
                                                    }
                                                )
                                    except:
                                        pass
                    except:
                        pass
            
            data = {
                'weight': weight,
                'entry_id': latest.entry_id if weight > 0 else '--',
                'created_at': latest.created_at.isoformat() if latest.created_at else None,
                'has_object': weight > 0,
                'latitude': float(latest.latitude) if latest.latitude else -6.152256,
                'longitude': float(latest.longitude) if latest.longitude else 106.694091,
                'source': 'database'
            }
            return JsonResponse({'status': 'success', 'data': data})
        
        # ✅ Jika tidak ada di database, coba dari ThingSpeak
        api_setting = APISetting.objects.first()
        if api_setting and api_setting.channel_id and api_setting.read_api_key:
            try:
                url = f"https://api.thingspeak.com/channels/{api_setting.channel_id}/feeds/last.json"
                params = {'api_key': api_setting.read_api_key}
                response = requests.get(url, params=params, timeout=5)
                
                if response.status_code == 200:
                    feed = response.json()
                    if feed:
                        weight = float(feed.get('field1', 0))
                        ts_created = feed.get('created_at')
                        
                        if ts_created:
                            try:
                                ts_time = datetime.strptime(ts_created, '%Y-%m-%dT%H:%M:%SZ')
                                ts_time = timezone.make_aware(ts_time)
                                ts_diff = (now - ts_time).total_seconds()
                                
                                # ✅ JIKA DATA LEBIH DARI 30 DETIK, ANGGAP KOSONG
                                if ts_diff > 30:
                                    weight = 0
                            except:
                                pass
                        
                        # Simpan ke database jika weight > 0
                        if weight > 0:
                            try:
                                created_at_str = feed.get('created_at')
                                if created_at_str:
                                    created_at = datetime.strptime(created_at_str, '%Y-%m-%dT%H:%M:%SZ')
                                    created_at = timezone.make_aware(created_at)
                                else:
                                    created_at = timezone.now()
                                
                                WeightData.objects.update_or_create(
                                    entry_id=str(feed.get('entry_id')),
                                    defaults={
                                        'weight': weight,
                                        'latitude': float(feed.get('field2', -6.152256)),
                                        'longitude': float(feed.get('field3', 106.694091)),
                                        'created_at': created_at
                                    }
                                )
                            except Exception as e:
                                logger.warning(f"Save to database failed: {e}")
                        
                        data = {
                            'weight': weight,
                            'entry_id': feed.get('entry_id', '--') if weight > 0 else '--',
                            'created_at': feed.get('created_at'),
                            'has_object': weight > 0,
                            'latitude': float(feed.get('field2', -6.152256)),
                            'longitude': float(feed.get('field3', 106.694091)),
                            'source': 'thingspeak'
                        }
                        return JsonResponse({'status': 'success', 'data': data})
            except Exception as e:
                logger.warning(f"ThingSpeak fallback error: {e}")
        
        # Default jika tidak ada data
        data = {
            'weight': 0.00,
            'entry_id': '--',
            'created_at': timezone.now().isoformat(),
            'has_object': False,
            'latitude': -6.152256,
            'longitude': 106.694091,
            'source': 'default'
        }
        
        return JsonResponse({'status': 'success', 'data': data})
        
    except Exception as e:
        logger.error(f"API latest error: {e}")
        return JsonResponse({
            'status': 'error',
            'message': str(e),
            'data': {
                'weight': 0.00,
                'entry_id': '--',
                'created_at': timezone.now().isoformat(),
                'has_object': False,
                'latitude': -6.152256,
                'longitude': 106.694091
            }
        }, status=500)

@login_required
def api_sync(request):
    """API sync data dari ThingSpeak ke database - SIMPAN JUGA WEIGHT=0"""
    try:
        api_setting = APISetting.objects.first()
        
        if not api_setting:
            return JsonResponse({
                'status': 'warning',
                'message': 'API settings not configured. Please go to Settings menu.'
            })
        
        if not api_setting.channel_id or not api_setting.read_api_key:
            return JsonResponse({
                'status': 'warning',
                'message': 'Channel ID or Read API Key is empty. Please configure in Settings.'
            })
        
        # Ambil data terakhir dari ThingSpeak
        url = f"https://api.thingspeak.com/channels/{api_setting.channel_id}/feeds/last.json"
        params = {'api_key': api_setting.read_api_key}
        
        response = requests.get(url, params=params, timeout=10)
        
        if response.status_code == 200:
            feed = response.json()
            
            if feed:
                entry_id = feed.get('entry_id')
                weight = float(feed.get('field1', 0))
                latitude = float(feed.get('field2', -6.152256))
                longitude = float(feed.get('field3', 106.694091))
                created_at_str = feed.get('created_at')
                
                if created_at_str:
                    try:
                        created_at = datetime.strptime(created_at_str, '%Y-%m-%dT%H:%M:%SZ')
                        created_at = timezone.make_aware(created_at)
                    except:
                        created_at = timezone.now()
                else:
                    created_at = timezone.now()
                
                # ✅ SIMPAN KE DATABASE (TERMASUK WEIGHT = 0)
                obj, created = WeightData.objects.update_or_create(
                    entry_id=str(entry_id) if entry_id else '0',
                    defaults={
                        'weight': weight,
                        'latitude': latitude,
                        'longitude': longitude,
                        'created_at': created_at
                    }
                )
                
                try:
                    UserActivity.objects.create(
                        user=request.user,
                        action='Sync from ThingSpeak',
                        details=f'Synced weight: {weight} kg, entry: {entry_id}',
                        ip_address=request.META.get('REMOTE_ADDR', '')
                    )
                except:
                    pass
                
                return JsonResponse({
                    'status': 'success',
                    'message': f'Sync successful. Weight: {weight} kg',
                    'latest_weight': weight,
                    'latest_entry_id': str(entry_id) if entry_id else '--',
                    'created': created
                })
            else:
                return JsonResponse({
                    'status': 'warning',
                    'message': 'No data from ThingSpeak channel'
                })
        else:
            return JsonResponse({
                'status': 'error',
                'message': f'ThingSpeak API error: HTTP {response.status_code}'
            })
            
    except requests.exceptions.Timeout:
        return JsonResponse({
            'status': 'error',
            'message': 'Connection timeout. Please check your internet connection.'
        }, status=408)
    except requests.exceptions.ConnectionError:
        return JsonResponse({
            'status': 'error',
            'message': 'Cannot connect to ThingSpeak. Check your internet connection.'
        }, status=503)
    except Exception as e:
        logger.error(f"API sync error: {e}", exc_info=True)
        return JsonResponse({
            'status': 'error',
            'message': f'Sync error: {str(e)}'
        }, status=500)
        
# ============================================
# USER MANAGEMENT - FDA COMPLIANT
# ============================================

@login_required
def user_management(request):
    """User management page with FDA Part 11 compliance"""
    users = User.objects.all()
    activities = UserActivity.objects.all().order_by('-timestamp')[:50]
    return render(request, 'dashboard/user_management.html', {
        'users': users,
        'activities': activities
    })

@login_required
def add_user(request):
    """Add new user - FDA compliant"""
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)
    
    try:
        # Cek permission
        if not request.user.is_superuser:
            return JsonResponse({'status': 'error', 'message': 'Only Super Admin can add users'}, status=403)
        
        username = request.POST.get('username', '').strip()
        email = request.POST.get('email', '').strip()
        password = request.POST.get('password', '')
        password2 = request.POST.get('password2', '')
        is_superuser = request.POST.get('is_superuser') == 'on'
        is_staff = request.POST.get('is_staff') == 'on'
        
        # Validasi
        if not username:
            return JsonResponse({'status': 'error', 'message': 'Username is required'})
        
        if User.objects.filter(username__iexact=username).exists():
            return JsonResponse({'status': 'error', 'message': 'Username already exists'})
        
        if len(password) < 4:
            return JsonResponse({'status': 'error', 'message': 'Password must be at least 4 characters'})
        
        if password != password2:
            return JsonResponse({'status': 'error', 'message': 'Passwords do not match'})
        
        # Buat user
        user = User.objects.create_user(
            username=username,
            email=email,
            password=password
        )
        
        if is_superuser:
            user.is_superuser = True
            user.is_staff = True
        elif is_staff:
            user.is_staff = True
        user.save()
        
        # Log aktivitas
        try:
            UserActivity.objects.create(
                user=request.user,
                action='Add User',
                details=f'Created user: {username} (Superuser: {is_superuser}, Staff: {is_staff})',
                ip_address=request.META.get('REMOTE_ADDR', '')
            )
        except Exception as e:
            print(f"Logging error: {e}")
        
        return JsonResponse({'status': 'success', 'message': f'User {username} created successfully'})
        
    except Exception as e:
        print(f"Add user error: {e}")
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

@login_required
def delete_user(request):
    """Delete user with FDA compliance"""
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)
    
    try:
        # Cek permission - hanya Super Admin yang bisa delete
        if not request.user.is_superuser:
            return JsonResponse({'status': 'error', 'message': 'Only Super Admin can delete users'}, status=403)
        
        data = json.loads(request.body)
        username = data.get('username')
        password = data.get('password')
        
        # Validasi password (FDA 21 CFR Part 11)
        if not password:
            return JsonResponse({'status': 'error', 'message': 'Password required for FDA compliance'}, status=400)
        
        if not request.user.check_password(password):
            return JsonResponse({'status': 'error', 'message': 'Wrong password'}, status=401)
        
        # Cari user yang akan dihapus
        user_to_delete = get_object_or_404(User, username__iexact=username)
        
        # Cegah delete sendiri
        if user_to_delete == request.user:
            return JsonResponse({'status': 'error', 'message': 'Cannot delete your own account'})
        
        # Cegah delete admin default
        if user_to_delete.username.lower() == 'admin':
            return JsonResponse({'status': 'error', 'message': 'Cannot delete default admin user'})
        
        # Log aktivitas sebelum delete
        try:
            UserActivity.objects.create(
                user=request.user,
                action='Delete User',
                details=f'Deleted user: {username}',
                ip_address=request.META.get('REMOTE_ADDR', '')
            )
        except:
            pass
        
        # Hapus user
        user_to_delete.delete()
        
        return JsonResponse({'status': 'success', 'message': f'User {username} deleted successfully'})
        
    except User.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'User not found'}, status=404)
    except json.JSONDecodeError:
        return JsonResponse({'status': 'error', 'message': 'Invalid JSON data'}, status=400)
    except Exception as e:
        print(f"Delete user error: {e}")
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
    
@login_required
def delete_all_users(request):
    """Delete all non-admin users with FDA compliance"""
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)
    
    try:
        # Cek permission
        if not request.user.is_superuser:
            return JsonResponse({'status': 'error', 'message': 'Only Super Admin can delete all users'}, status=403)
        
        data = json.loads(request.body)
        password = data.get('password')
        confirm = data.get('confirm')
        
        if not password:
            return JsonResponse({'status': 'error', 'message': 'Password required for FDA compliance'}, status=400)
        
        if not request.user.check_password(password):
            return JsonResponse({'status': 'error', 'message': 'Wrong password'}, status=401)
        
        if confirm != 'DELETE_ALL':
            return JsonResponse({'status': 'error', 'message': 'Confirmation text must be "DELETE_ALL"'}, status=400)
        
        # Hapus semua user kecuali admin dan diri sendiri
        users_to_delete = User.objects.exclude(username__iexact='admin').exclude(id=request.user.id)
        count = users_to_delete.count()
        
        # Log sebelum delete
        try:
            UserActivity.objects.create(
                user=request.user,
                action='Delete All Users',
                details=f'Deleted {count} users',
                ip_address=request.META.get('REMOTE_ADDR', '')
            )
        except:
            pass
        
        users_to_delete.delete()
        
        return JsonResponse({'status': 'success', 'message': f'{count} users deleted successfully'})
        
    except json.JSONDecodeError:
        return JsonResponse({'status': 'error', 'message': 'Invalid JSON data'}, status=400)
    except Exception as e:
        print(f"Delete all users error: {e}")
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
    
@login_required
def export_users_excel(request):
    """Export users to Excel with FDA compliance"""
    try:
        password = request.GET.get('password')
        if not password:
            return JsonResponse({
                'status': 'error', 
                'message': 'Password required for FDA Part 11 compliance'
            }, status=401)
        
        if not request.user.check_password(password):
            return JsonResponse({
                'status': 'error', 
                'message': 'Invalid password'
            }, status=401)
        
        users = User.objects.all().values('username', 'email', 'date_joined', 'last_login', 'is_active', 'is_superuser', 'is_staff')
        df = pd.DataFrame(list(users))
        
        if not df.empty:
            df['date_joined'] = pd.to_datetime(df['date_joined'])
            df['last_login'] = pd.to_datetime(df['last_login'])
            df['role'] = df.apply(lambda row: 'Super Admin' if row['is_superuser'] else ('Staff' if row['is_staff'] else 'User'), axis=1)
        
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Users', index=False)
            
            workbook = writer.book
            worksheet = writer.sheets['Users']
            
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            header_font = Font(bold=True, size=11, color='FFFFFF')
            header_fill = PatternFill(start_color='0D47A1', end_color='0D47A1', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center')
            
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            metadata = pd.DataFrame({
                'Report Info': ['Generated By', 'Generated At', 'Total Users', 'FDA Compliant'],
                'Value': [
                    request.user.username,
                    timezone.now().strftime('%Y-%m-%d %H:%M:%S'),
                    len(users),
                    'Yes - 21 CFR Part 11'
                ]
            })
            metadata.to_excel(writer, sheet_name='Metadata', index=False)
        
        try:
            UserActivity.objects.create(
                user=request.user,
                action='Export Users',
                details=f'Exported {len(users)} users to Excel',
                ip_address=request.META.get('REMOTE_ADDR')
            )
        except:
            pass
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'users_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
        
    except Exception as e:
        logger.error(f"Export users error: {e}", exc_info=True)
        return JsonResponse({
            'status': 'error',
            'message': f'Export failed: {str(e)}'
        }, status=500)

@login_required
def export_audit_csv(request):
    """Export audit log to CSV with FDA compliance"""
    try:
        activities = UserActivity.objects.all().order_by('-timestamp')
        
        response = HttpResponse(content_type='text/csv')
        filename = f'audit_log_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        writer = csv.writer(response)
        writer.writerow(['Timestamp', 'User', 'Action', 'Details', 'IP Address'])
        
        for item in activities:
            writer.writerow([
                item.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
                item.user.username,
                item.action,
                item.details or '-',
                item.ip_address or '-'
            ])
        
        try:
            UserActivity.objects.create(
                user=request.user,
                action='Export Audit Log',
                details=f'Exported {len(activities)} audit records to CSV',
                ip_address=request.META.get('REMOTE_ADDR')
            )
        except:
            pass
        
        return response
        
    except Exception as e:
        logger.error(f"Export audit error: {e}", exc_info=True)
        return JsonResponse({
            'status': 'error',
            'message': f'Export failed: {str(e)}'
        }, status=500)

@login_required
def delete_audit_log(request):
    """Delete single audit log (Super Admin only)"""
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Method not allowed'}, status=405)
    
    try:
        if not request.user.is_superuser:
            return JsonResponse({'status': 'error', 'message': 'Only Super Admin can delete audit logs'}, status=403)
        
        data = json.loads(request.body)
        log_id = data.get('id')
        password = data.get('password')
        
        if not password:
            return JsonResponse({'status': 'error', 'message': 'Password required for FDA compliance'}, status=400)
        
        if not request.user.check_password(password):
            return JsonResponse({'status': 'error', 'message': 'Wrong password'}, status=401)
        
        if not log_id:
            return JsonResponse({'status': 'error', 'message': 'Log ID is required'}, status=400)
        
        log = get_object_or_404(UserActivity, id=log_id)
        
        # Log aktivitas sebelum delete
        try:
            UserActivity.objects.create(
                user=request.user,
                action='Delete Audit Log',
                details=f'Deleted audit log ID: {log_id}',
                ip_address=request.META.get('REMOTE_ADDR', '')
            )
        except:
            pass
        
        log.delete()
        
        return JsonResponse({'status': 'success', 'message': 'Audit log deleted successfully'})
        
    except UserActivity.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Audit log not found'}, status=404)
    except json.JSONDecodeError:
        return JsonResponse({'status': 'error', 'message': 'Invalid JSON data'}, status=400)
    except Exception as e:
        print(f"Delete audit log error: {e}")
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
    
@login_required
def delete_all_audit_logs(request):
    """Delete all audit logs (Super Admin only)"""
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Method not allowed'}, status=405)
    
    try:
        if not request.user.is_superuser:
            return JsonResponse({'status': 'error', 'message': 'Only Super Admin can delete audit logs'}, status=403)
        
        data = json.loads(request.body)
        password = data.get('password')
        confirm = data.get('confirm')
        
        if not password:
            return JsonResponse({'status': 'error', 'message': 'Password required for FDA compliance'}, status=400)
        
        if not request.user.check_password(password):
            return JsonResponse({'status': 'error', 'message': 'Wrong password'}, status=401)
        
        if confirm != 'DELETE_ALL_AUDIT':
            return JsonResponse({'status': 'error', 'message': 'Confirmation text must be "DELETE_ALL_AUDIT"'}, status=400)
        
        count = UserActivity.objects.count()
        
        # Log sebelum delete
        try:
            UserActivity.objects.create(
                user=request.user,
                action='Delete All Audit Logs',
                details=f'Deleted ALL {count} audit logs',
                ip_address=request.META.get('REMOTE_ADDR', '')
            )
        except:
            pass
        
        UserActivity.objects.all().delete()
        
        return JsonResponse({'status': 'success', 'message': f'All {count} audit logs deleted successfully'})
        
    except json.JSONDecodeError:
        return JsonResponse({'status': 'error', 'message': 'Invalid JSON data'}, status=400)
    except Exception as e:
        print(f"Delete all audit logs error: {e}")
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)